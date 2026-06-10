#!/usr/bin/env python3
"""Lending P&L HTML Report — Cake.

Build single HTML report 3 sections:
  1. Overall (Disb/TOI/PBO/PBT × 12 month × 3 period)
  2. By Product (CashLoan / Overdraft / Payday / Paylater)
  3. By Channel × Product matrix (2025 Actual FY)

Output: /tmp/lending-pl-report.html (mở bằng browser)
Usage: python3 lending_pl_html_report.py
"""
import openpyxl
from pathlib import Path
import html

REPO_ROOT = Path(__file__).resolve().parents[2]
RAW = REPO_ROOT / "Raw" / "Finance" / "P&L"
OUT = Path("/tmp/lending-pl-report.html")

MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

# 2026 Actual is closed only through this many months (rest in source file = partial/draft).
# AUTO-DETECTED at runtime in main() from the disbursement anchor — do NOT hand-edit.
# Period dict-keys are month-agnostic ("2026 YTD") so they never change month-to-month;
# only this count + the display label (YTD_LABEL) + annualization (ANNUALIZE) move.
ACTUAL_2026_COMPLETE_MONTHS = 4   # placeholder; overwritten by detect_complete_months()
YTD_LABEL = "YTD-Apr"             # display only; refined to "YTD-<Mon>" in main()
ANNUALIZE = 3.0                   # 12 / complete_months; recomputed in main()


def num(v, div=1.0):
    if v is None or isinstance(v, str): return None
    return v / div


def safe_sum(vals):
    pool = [v for v in vals if v is not None]
    return sum(pool) if pool else None


# ========== OVERALL ==========
# (file, sheet, row, label_col, expected_label, month_start, month_end, unit_divisor)
OVERALL_CFG = {
    "2025 Actual": {"path": RAW/"P&L_2025_final.xlsx", "sheet": "Lending",
                    "month_start": 16, "month_end": 27, "div": 1e9,
                    "rows": {"disb": 7, "toi": 19, "prov": 30, "pbo": 46, "pbt": 45}, "label_col": 12},
    "2026 Actual": {"path": RAW/"P&L_2026_Actual.xlsx", "sheet": "Total",
                    "month_start": 6, "month_end": 17, "div": 1e9,
                    "rows": {"disb": 6, "toi": 17, "prov": 36, "pbo": 52, "pbt": 53}, "label_col": 2},
    "2026 Budget": {"path": RAW/"P&L-Budget-2026.xlsx", "sheet": "Lending",
                    "month_start": 62, "month_end": 73, "div": 1.0,
                    "rows": {"disb": 7, "toi": 11, "prov": 28, "pbo": 45, "pbt": 47}, "label_col": 6},
}

# Provision sign normalization: source files use mixed conventions
# (2025 stores +, 2026 stores −). Always normalize to ABS value = expense magnitude.
PROV_METRICS = {"prov"}


def detect_complete_months():
    """# of closed 2026 months = count of leading non-empty disbursement cells.

    Disb is the anchor: it stays blank (' - ' → None) until a month fully closes,
    whereas TOI/PBT can leak a draft value a month early. Stops at the first gap so
    a stray future value never inflates the count.
    """
    cfg = OVERALL_CFG["2026 Actual"]
    wb = openpyxl.load_workbook(cfg["path"], data_only=True)
    ws = wb[cfg["sheet"]]
    row = cfg["rows"]["disb"]
    n = 0
    for c in range(cfg["month_start"], cfg["month_end"] + 1):
        if num(ws.cell(row, c).value, cfg["div"]) is None:
            break
        n += 1
    wb.close()
    return n

def load_overall():
    out = {}
    for period, cfg in OVERALL_CFG.items():
        wb = openpyxl.load_workbook(cfg["path"], data_only=True)
        ws = wb[cfg["sheet"]]
        out[period] = {}
        for metric, row in cfg["rows"].items():
            vals = [num(ws.cell(row, c).value, cfg["div"]) for c in range(cfg["month_start"], cfg["month_end"]+1)]
            if metric in PROV_METRICS:
                vals = [abs(v) if v is not None else None for v in vals]
            if period == "2026 Actual":
                vals = vals[:ACTUAL_2026_COMPLETE_MONTHS] + [None] * (12 - ACTUAL_2026_COMPLETE_MONTHS)
            out[period][metric] = vals
        wb.close()
    return out


# ========== BY PRODUCT ==========
# 2026 Actual: row positions in each product sheet (disb, toi, pbo) — top-level (R6 = product all)
ACTUAL_2026_PRODUCT = {
    "CashLoan": {"sheet": "CL", "rows": {"disb": 6, "toi": 19, "pbo": 40}},
    "Payday":   {"sheet": "Payday", "rows": {"disb": 6, "toi": 19, "pbo": 40}},
    "Overdraft":{"sheet": "OD", "rows": {"disb": 6, "toi": 20, "pbo": 41}},
    "Paylater": {"sheet": "Paylater", "rows": {"disb": 8, "toi": 25, "pbo": 52}},  # disb = R8 Spending (BNPL)
}

# 2026 Budget Lending sheet — top-level product rows
BUDGET_2026_PRODUCT = {
    "CashLoan":  {"disb": 55, "toi": 59, "pbo": 88},
    "Payday":    {"disb": 278, "toi": 282, "pbo": 311},
    "Overdraft": {"disb": 426, "toi": 432, "pbo": 460},
    "Paylater":  {"disb": 469, "toi": 474, "pbo": 503},
}

# 2025 product: sum from product×channel rows. Each product's channel rows in 2025 Lending sheet:
# Cashloan: VDSCL R52, VDSPRCL R91, BEGCL R130, NGSCL R169, CAKECL R208, MWGCL R247,
#           VNPAYCL R286, ZALOPAYCL R325, VNPOSTCL R364, VNPOSTPRCL R403, MISACL R442
# Overdraft: CAKEOD R481
# Payday: VDSPD R520, VDSPRPD R559, CAKEPD R598, VNPAYPD R637, ZALOPAYPD R675
# Each section: Disbursement at base, TOI at +12, Profit before allocation cost at +33
ACTUAL_2025_PRODUCT_CHANNEL_BASES = {
    "CashLoan": [52, 91, 130, 169, 208, 247, 286, 325, 364, 403, 442],
    "Overdraft": [481],
    "Payday": [520, 559, 598, 637, 675],
    "Paylater": [],  # not in 2025 Lending sheet by this naming
}
CHANNEL_NAMES_2025 = {
    52: "VDS",  91: "VDS-PR", 130: "BEG", 169: "NGS", 208: "CAKE", 247: "MWG",
    286: "VNPAY", 325: "ZALOPAY", 364: "VNPOST", 403: "VNPOST-PR", 442: "MISA",
    481: "CAKE",
    520: "VDS", 559: "VDS-PR", 598: "CAKE", 637: "VNPAY", 675: "ZALOPAY",
}
# offset within each section: row = base + offset
PRODUCT_CHANNEL_OFFSETS = {"disb": 0, "toi": 12, "pbo": 33}


def load_by_product():
    """Returns {product: {metric: {period: monthly_list or None}}}."""
    result = {p: {"disb": {}, "toi": {}, "pbo": {}} for p in ACTUAL_2026_PRODUCT}

    # 2026 Actual — clip to closed months only
    wb = openpyxl.load_workbook(RAW/"P&L_2026_Actual.xlsx", data_only=True)
    for product, cfg in ACTUAL_2026_PRODUCT.items():
        ws = wb[cfg["sheet"]]
        for metric, row in cfg["rows"].items():
            if row is None:
                result[product][metric]["2026 Actual"] = [None]*12
                continue
            vals = [num(ws.cell(row, c).value, 1e9) for c in range(6, 18)]
            vals = vals[:ACTUAL_2026_COMPLETE_MONTHS] + [None] * (12 - ACTUAL_2026_COMPLETE_MONTHS)
            result[product][metric]["2026 Actual"] = vals
    wb.close()

    # 2026 Budget
    wb = openpyxl.load_workbook(RAW/"P&L-Budget-2026.xlsx", data_only=True)
    ws = wb["Lending"]
    for product, rows in BUDGET_2026_PRODUCT.items():
        for metric, row in rows.items():
            vals = [num(ws.cell(row, c).value, 1.0) for c in range(62, 74)]
            result[product][metric]["2026 Budget"] = vals
    wb.close()

    # 2025 Actual — sum product×channel
    wb = openpyxl.load_workbook(RAW/"P&L_2025_final.xlsx", data_only=True)
    ws = wb["Lending"]
    for product, bases in ACTUAL_2025_PRODUCT_CHANNEL_BASES.items():
        for metric in ["disb", "toi", "pbo"]:
            monthly = [None]*12
            for m in range(12):
                month_col = 16 + m
                vals = []
                for base in bases:
                    row = base + PRODUCT_CHANNEL_OFFSETS[metric]
                    v = num(ws.cell(row, month_col).value, 1e9)
                    if v is not None:
                        vals.append(v)
                monthly[m] = sum(vals) if vals else None
            result[product][metric]["2025 Actual"] = monthly
    wb.close()

    return result


# ========== CHANNEL DEEP DIVE ==========
# Grouping (theo Đạm): CAKE / VDS / ZLP / MWG / Others. VDS-PR merge vào VDS sau extraction.
# Mapping channel → (product → row base) for 2025 + 2026. Row offsets verified vs 2025 pattern.
CHANNEL_DEEP = {
    "VDS": {
        "2025": {"CL": 52, "Payday": 520},
        "2026": {"CL": 47, "Payday": 46},
    },
    "VDS-PR": {  # merged into VDS post-extraction
        "2025": {"CL": 91, "Payday": 559},
        "2026": {"CL": 87, "Payday": 86},
    },
    "CAKE": {
        "2025": {"CL": 208, "OD": 481, "Payday": 598},
        "2026": {"CL": 207, "OD": 6, "Payday": 126},
    },
    "ZLP": {  # ZALOPAY — CL + Payday
        "2025": {"CL": 325, "Payday": 675},
        "2026": {"CL": 327, "Payday": 207},
    },
    "MWG": {  # CL only (no Payday/OD channel for MWG)
        "2025": {"CL": 247},
        "2026": {"CL": 247},
    },
}
BUDGET_CHANNEL_ROWS = {
    "CAKE": {
        "CL":     {"disb": 94,  "toi": 98,  "prov": 109, "pbo": 123},
        "Payday": {"disb": 318, "toi": 322, "prov": 332, "pbo": 346},
        "OD":     {"disb": 426, "toi": 432, "prov": 449, "pbo": 460},
    },
    "VDS": {
        "CL":     {"disb": 166, "toi": 170, "prov": 181, "pbo": 195},
        "Payday": {"disb": 353, "toi": 357, "prov": 367, "pbo": 381},
    },
}
# CAKE CL xsale sub-line (R127-161) — must be summed with CAKE CL organic above
BUDGET_CHANNEL_XSALE_ROWS = {
    "CAKE": {
        "CL": {"disb": 130, "toi": 134, "prov": 145, "pbo": 159},
    },
}
OFFSET_2025 = {"disb": 0, "toi": 12, "prov": 23, "pbo": 33}
OFFSET_2026_CLPD = {"disb": 0, "toi": 13, "prov": 24, "pbo": 34}
OFFSET_2026_OD = {"disb": 0, "toi": 14, "prov": 25, "pbo": 35}

# Risk metric offsets (state metrics — use end-of-period column only)
# 2026 channel sub-section: from base (Disb row), LG2 +10, NPL CAKE +11, NPL CIC +12, Balance ENR +3
OFFSET_2026_CHANNEL_RISK = {"lg2": 10, "npl_cake": 11, "npl_cic": 12, "balance": 3}
# 2025 channel sub-section: NPL CAKE +10, NPL CIC +11, LG2 numerator (Overdue 11-90d) +9, Balance ENR +3
OFFSET_2025_CHANNEL_RISK = {"npl_cake_pct": 10, "npl_cic_pct": 11, "lg2_overdue": 9, "balance": 3}


def load_channel_deep():
    """Returns {channel: {period: {product: {metric: value (Tỷ VND)}}}}.
    period = '2025 FY' | '2026 YTD' | '2026 Budget YTD' | '2026 Budget FY'."""
    result = {}

    wb25 = openpyxl.load_workbook(RAW/"P&L_2025_final.xlsx", data_only=True)
    ws25 = wb25["Lending"]
    wb26 = openpyxl.load_workbook(RAW/"P&L_2026_Actual.xlsx", data_only=True)
    wbB = openpyxl.load_workbook(RAW/"P&L-Budget-2026.xlsx", data_only=True)
    wsB = wbB["Lending"]

    def absify(metric, vals):
        if metric in PROV_METRICS:
            return [abs(v) if v is not None else None for v in vals]
        return vals

    for ch, periods in CHANNEL_DEEP.items():
        result[ch] = {"2025 FY": {}, "2026 YTD": {}, "2026 Budget YTD": {}, "2026 Budget FY": {}}
        for product, base in periods["2025"].items():
            result[ch]["2025 FY"][product] = {}
            for metric, off in OFFSET_2025.items():
                row = base + off
                vals = [num(ws25.cell(row, c).value, 1e9) for c in range(16, 28)]
                vals = absify(metric, vals)
                result[ch]["2025 FY"][product][metric] = safe_sum(vals)
            # 2025 channel risk (state at FY end = Dec 25 col 27): NPL CAKE + CIC, LG2 % (computed)
            bal = num(ws25.cell(base + OFFSET_2025_CHANNEL_RISK["balance"], 27).value, 1e9)
            npl_cake = ws25.cell(base + OFFSET_2025_CHANNEL_RISK["npl_cake_pct"], 27).value
            npl_cic = ws25.cell(base + OFFSET_2025_CHANNEL_RISK["npl_cic_pct"], 27).value
            lg2_overdue = num(ws25.cell(base + OFFSET_2025_CHANNEL_RISK["lg2_overdue"], 27).value, 1e9)
            result[ch]["2025 FY"][product]["balance"] = bal
            result[ch]["2025 FY"][product]["npl_cake"] = npl_cake if isinstance(npl_cake, (int, float)) else None
            result[ch]["2025 FY"][product]["npl_cic"] = npl_cic if isinstance(npl_cic, (int, float)) else None
            result[ch]["2025 FY"][product]["lg2"] = (lg2_overdue / bal) if (bal and lg2_overdue is not None) else None
        for product, base in periods["2026"].items():
            ws = wb26[product]
            offsets = OFFSET_2026_OD if product == "OD" else OFFSET_2026_CLPD
            result[ch]["2026 YTD"][product] = {}
            for metric, off in offsets.items():
                row = base + off
                vals = [num(ws.cell(row, c).value, 1e9) for c in range(6, 6 + ACTUAL_2026_COMPLETE_MONTHS)]
                vals = absify(metric, vals)
                result[ch]["2026 YTD"][product][metric] = safe_sum(vals)
            # 2026 channel risk = state at the last closed month. Col 6 = Jan, so month m → col 5+m.
            def _numcell(cell_v):
                return cell_v if isinstance(cell_v, (int, float)) else None
            end_col = 5 + ACTUAL_2026_COMPLETE_MONTHS  # e.g. Apr→9, May→10
            if product != "OD":
                bal = num(ws.cell(base + OFFSET_2026_CHANNEL_RISK["balance"], end_col).value, 1e9)
                result[ch]["2026 YTD"][product]["balance"] = bal
                result[ch]["2026 YTD"][product]["lg2"] = _numcell(ws.cell(base + OFFSET_2026_CHANNEL_RISK["lg2"], end_col).value)
                result[ch]["2026 YTD"][product]["npl_cake"] = _numcell(ws.cell(base + OFFSET_2026_CHANNEL_RISK["npl_cake"], end_col).value)
                result[ch]["2026 YTD"][product]["npl_cic"] = _numcell(ws.cell(base + OFFSET_2026_CHANNEL_RISK["npl_cic"], end_col).value)
            else:
                # OD = single channel (CAKE), top-level rows R9 Balance, R17 LG2, R18 NPL CAKE, R19 NPL CIC
                result[ch]["2026 YTD"][product]["balance"] = num(ws.cell(9, end_col).value, 1e9)
                result[ch]["2026 YTD"][product]["lg2"] = _numcell(ws.cell(17, end_col).value)
                result[ch]["2026 YTD"][product]["npl_cake"] = _numcell(ws.cell(18, end_col).value)
                result[ch]["2026 YTD"][product]["npl_cic"] = _numcell(ws.cell(19, end_col).value)
        if ch in BUDGET_CHANNEL_ROWS:
            for product, rows in BUDGET_CHANNEL_ROWS[ch].items():
                result[ch]["2026 Budget YTD"][product] = {}
                result[ch]["2026 Budget FY"][product] = {}
                for metric, row in rows.items():
                    vals = [num(wsB.cell(row, c).value, 1.0) for c in range(62, 74)]
                    # Add xsale sub-line if exists (e.g. CAKE CL xsale)
                    xsale_rows = BUDGET_CHANNEL_XSALE_ROWS.get(ch, {}).get(product, {})
                    if metric in xsale_rows:
                        xs_vals = [num(wsB.cell(xsale_rows[metric], c).value, 1.0) for c in range(62, 74)]
                        vals = [(a or 0) + (b or 0) if (a is not None or b is not None) else None
                                for a, b in zip(vals, xs_vals)]
                    vals = absify(metric, vals)
                    result[ch]["2026 Budget YTD"][product][metric] = safe_sum(vals[:ACTUAL_2026_COMPLETE_MONTHS])
                    result[ch]["2026 Budget FY"][product][metric] = safe_sum(vals)
    wb25.close()
    wb26.close()
    wbB.close()

    # ===== Merge VDS-PR vào VDS (theo Đạm: VDS-PR là sub-line của VDS family) =====
    # Flow metrics (disb/toi/prov/pbo) = sum. State balance = sum. Rate metrics (npl/lg2) = balance-weighted avg.
    FLOW_KEYS = {"disb", "toi", "prov", "pbo"}
    RATE_KEYS = {"npl_cake", "npl_cic", "lg2"}
    if "VDS-PR" in result and "VDS" in result:
        for period in result["VDS-PR"]:
            for product, pr_metrics in result["VDS-PR"][period].items():
                vds_period = result["VDS"].setdefault(period, {})
                if product not in vds_period:
                    vds_period[product] = dict(pr_metrics)
                    continue
                vds_m = vds_period[product]
                bal_vds = vds_m.get("balance") or 0
                bal_pr = pr_metrics.get("balance") or 0
                bal_total = bal_vds + bal_pr
                # Flow + balance: sum
                for k in FLOW_KEYS | {"balance"}:
                    a, b = vds_m.get(k), pr_metrics.get(k)
                    if a is not None or b is not None:
                        vds_m[k] = (a or 0) + (b or 0)
                # Rates: balance-weighted average
                for k in RATE_KEYS:
                    rv, rp = vds_m.get(k), pr_metrics.get(k)
                    if bal_total > 0 and (rv is not None or rp is not None):
                        vds_m[k] = ((rv or 0) * bal_vds + (rp or 0) * bal_pr) / bal_total
                    elif rv is None:
                        vds_m[k] = rp
        del result["VDS-PR"]

    return result


GROUP_CHANNELS = ["CAKE", "VDS", "ZLP", "MWG"]  # named groups; phần còn lại = Others


def load_channel_budget_monthly():
    """Returns {channel: {product: {metric: [12 monthly values]}}} for CAKE + VDS.
    Actual 2026 monthly (Jan-Apr closed, rest None) + Budget 2026 monthly (full 12).
    VDS-PR merged into VDS."""
    result = {}

    wb26 = openpyxl.load_workbook(RAW/"P&L_2026_Actual.xlsx", data_only=True)
    wbB = openpyxl.load_workbook(RAW/"P&L-Budget-2026.xlsx", data_only=True)
    wsB = wbB["Lending"]

    CHANNEL_ACTUAL_MAP = {
        "CAKE": {
            "CL":     {"sheet": "CL",     "base": 207, "offsets": OFFSET_2026_CLPD},
            "Payday": {"sheet": "Payday", "base": 126, "offsets": OFFSET_2026_CLPD},
            "OD":     {"sheet": "OD",     "base": 6,   "offsets": OFFSET_2026_OD},
        },
        "VDS": {
            "CL":     {"sheet": "CL",     "base": 47,  "offsets": OFFSET_2026_CLPD},
            "Payday": {"sheet": "Payday", "base": 46,  "offsets": OFFSET_2026_CLPD},
        },
        "VDS-PR": {
            "CL":     {"sheet": "CL",     "base": 87,  "offsets": OFFSET_2026_CLPD},
            "Payday": {"sheet": "Payday", "base": 86,  "offsets": OFFSET_2026_CLPD},
        },
    }

    METRICS_MONTHLY = ["disb", "toi", "prov", "pbo"]

    for ch, products in CHANNEL_ACTUAL_MAP.items():
        result[ch] = {}
        for product, cfg in products.items():
            ws = wb26[cfg["sheet"]]
            actual = {}
            for metric in METRICS_MONTHLY:
                row = cfg["base"] + cfg["offsets"][metric]
                vals = [num(ws.cell(row, c).value, 1e9) for c in range(6, 18)]
                if metric == "prov":
                    vals = [abs(v) if v is not None else None for v in vals]
                vals = vals[:ACTUAL_2026_COMPLETE_MONTHS] + [None] * (12 - ACTUAL_2026_COMPLETE_MONTHS)
                actual[metric] = vals
            result[ch][product] = {"actual_monthly": actual}

    # Budget monthly (full 12 months)
    for ch, products in BUDGET_CHANNEL_ROWS.items():
        for product, rows in products.items():
            budget = {}
            for metric, row in rows.items():
                vals = [num(wsB.cell(row, c).value, 1.0) for c in range(62, 74)]
                # Add xsale sub-line if exists (e.g. CAKE CL xsale)
                xsale_rows = BUDGET_CHANNEL_XSALE_ROWS.get(ch, {}).get(product, {})
                if metric in xsale_rows:
                    xs_vals = [num(wsB.cell(xsale_rows[metric], c).value, 1.0) for c in range(62, 74)]
                    vals = [(a or 0) + (b or 0) if (a is not None or b is not None) else None
                            for a, b in zip(vals, xs_vals)]
                if metric == "prov":
                    vals = [abs(v) if v is not None else None for v in vals]
                budget[metric] = vals
            if ch in result and product in result[ch]:
                result[ch][product]["budget_monthly"] = budget
            else:
                result.setdefault(ch, {}).setdefault(product, {})["budget_monthly"] = budget

    wb26.close()
    wbB.close()

    # Merge VDS-PR into VDS (monthly arrays element-wise sum)
    if "VDS-PR" in result:
        for product, data in result["VDS-PR"].items():
            if product in result.get("VDS", {}):
                for period_key in ["actual_monthly"]:
                    if period_key in data:
                        for metric in METRICS_MONTHLY:
                            vds_arr = result["VDS"][product][period_key].get(metric, [None]*12)
                            pr_arr = data[period_key].get(metric, [None]*12)
                            merged = []
                            for a, b in zip(vds_arr, pr_arr):
                                if a is not None or b is not None:
                                    merged.append((a or 0) + (b or 0))
                                else:
                                    merged.append(None)
                            result["VDS"][product][period_key][metric] = merged
        del result["VDS-PR"]

    return result


def render_channel_monthly_budget(ch_monthly):
    """Render monthly Actual vs Budget tables for CAKE + VDS channels."""
    parts = []
    parts.append("<h2>2.4 — Monthly Actual vs Budget by Channel × Product (CAKE + VDS)</h2>")
    parts.append("<p class='note'>Budget monthly chỉ map riêng cho <strong>CAKE</strong> + <strong>VDS</strong> (incl. VDS-PR). "
                 "Metrics: Disb, TOI, Provision, PBO. Đơn vị: Tỷ VND. "
                 "2026 Actual closed qua tháng " + str(ACTUAL_2026_COMPLETE_MONTHS) + " — còn lại = trống.</p>")

    METRIC_LABELS = {"disb": "Disbursement", "toi": "TOI", "prov": "Provision", "pbo": "PBO"}
    CH_LABELS = {"CAKE": "CAKE", "VDS": "VDS (incl. VDS-PR)"}

    for ch in ["CAKE", "VDS"]:
        if ch not in ch_monthly:
            continue
        parts.append(f"<h3>{CH_LABELS[ch]} — Monthly Actual vs Budget 2026</h3>")
        for product in ["CL", "Payday", "OD"]:
            if product not in ch_monthly[ch]:
                continue
            data = ch_monthly[ch][product]
            actual_m = data.get("actual_monthly", {})
            budget_m = data.get("budget_monthly", {})
            parts.append(f"<h4>{ch} × {product}</h4>")
            parts.append("<table class='data'><thead><tr><th>Metric</th><th>Period</th>")
            for m in MONTHS:
                parts.append(f"<th>{m}</th>")
            parts.append("<th class='accent'>YTD</th><th class='accent'>FY</th></tr></thead><tbody>")

            for metric in ["disb", "toi", "prov", "pbo"]:
                a_vals = actual_m.get(metric, [None]*12)
                b_vals = budget_m.get(metric, [None]*12)
                a_ytd = safe_sum(a_vals[:ACTUAL_2026_COMPLETE_MONTHS])
                b_ytd = safe_sum(b_vals[:ACTUAL_2026_COMPLETE_MONTHS])
                b_fy = safe_sum(b_vals)
                # Actual row
                parts.append(f"<tr><td class='label' rowspan='3'>{METRIC_LABELS[metric]}</td>"
                             f"<td class='label'>Actual</td>")
                for v in a_vals:
                    parts.append(f"<td class='{cell_class(v)}'>{fmt(v)}</td>")
                parts.append(f"<td class='accent'>{fmt(a_ytd)}</td><td class='accent'>-</td></tr>")
                # Budget row
                parts.append(f"<tr><td class='label'>Budget</td>")
                for v in b_vals:
                    parts.append(f"<td>{fmt(v)}</td>")
                parts.append(f"<td class='accent'>{fmt(b_ytd)}</td><td class='accent'>{fmt(b_fy)}</td></tr>")
                # Delta row
                is_prov = (metric == "prov")
                parts.append(f"<tr class='delta'><td class='label'>Δ</td>")
                for a, b in zip(a_vals, b_vals):
                    if a is not None and b is not None:
                        d = a - b
                        # For provision: positive delta = overshoot = bad
                        if is_prov:
                            cls = "neg" if d > 0 else ("pos" if d < 0 else "")
                        else:
                            cls = cell_class(d, neg_bad=False)
                        parts.append(f"<td class='{cls}'>{fmt(d)}</td>")
                    else:
                        parts.append("<td>-</td>")
                # YTD delta
                if a_ytd is not None and b_ytd is not None:
                    d_ytd = a_ytd - b_ytd
                    pct_ytd = d_ytd / b_ytd * 100 if b_ytd else 0
                    if is_prov:
                        cls_ytd = "neg" if d_ytd > 0 else ("pos" if d_ytd < 0 else "")
                    else:
                        cls_ytd = cell_class(d_ytd, neg_bad=False)
                    parts.append(f"<td class='accent {cls_ytd}'>{fmt(d_ytd)} ({pct_ytd:+.0f}%)</td>")
                else:
                    parts.append("<td class='accent'>-</td>")
                parts.append("<td class='accent'>-</td></tr>")

            parts.append("</tbody></table>")

    return "\n".join(parts)


def render_channel_deep_dive(deep, overall):
    parts = []
    parts.append("<p class='note'>Tách kênh theo Đạm: <strong>CAKE</strong> (direct app) · <strong>VDS</strong> (Viettel — đã merge VDS-PR) · <strong>ZLP</strong> (ZaloPay) · <strong>MWG</strong> (Mobile World) · <strong>Others</strong> (= Overall − 4 kênh trên: BEG, NGS, VNPAY, VNPOST, MISA, leadgen, xsale, IR-tier, Paylater channels). Budget chỉ map được CAKE + VDS — ZLP/MWG/Others budget hiển thị '-' (Others budget = Overall − CAKE − VDS).</p>")

    def channel_sum(ch, period, metric):
        """Sum across all products for a channel in given period."""
        if ch not in deep or period not in deep[ch]: return None
        vals = [deep[ch][period][p].get(metric) for p in deep[ch][period]]
        vals = [v for v in vals if v is not None]
        return sum(vals) if vals else None

    METRICS_4 = ["disb", "toi", "prov", "pbo"]
    # Per-group actual / budget / budget-FY
    grp_a  = {ch: {m: channel_sum(ch, "2026 YTD", m) for m in METRICS_4} for ch in GROUP_CHANNELS}
    grp_b  = {ch: {m: channel_sum(ch, "2026 Budget YTD", m) for m in METRICS_4} for ch in GROUP_CHANNELS}
    grp_bf = {ch: {m: channel_sum(ch, "2026 Budget FY", m) for m in METRICS_4} for ch in GROUP_CHANNELS}
    # Overall
    ov_a  = {m: safe_sum(overall["2026 Actual"][m][:ACTUAL_2026_COMPLETE_MONTHS]) for m in METRICS_4}
    ov_b  = {m: safe_sum(overall["2026 Budget"][m][:ACTUAL_2026_COMPLETE_MONTHS]) for m in METRICS_4}
    ov_bf = {m: safe_sum(overall["2026 Budget"][m]) for m in METRICS_4}
    # Others = Overall − named groups (budget: only CAKE+VDS have budget → Others_b = Overall − CAKE − VDS)
    oth_a  = {m: ov_a[m]  - sum((grp_a[ch][m]  or 0) for ch in GROUP_CHANNELS) for m in METRICS_4}
    oth_b  = {m: ov_b[m]  - sum((grp_b[ch][m]  or 0) for ch in GROUP_CHANNELS) for m in METRICS_4}
    oth_bf = {m: ov_bf[m] - sum((grp_bf[ch][m] or 0) for ch in GROUP_CHANNELS) for m in METRICS_4}

    parts.append(f"<h3>4.1 — Channel Split Summary ({YTD_LABEL} 2026, Tỷ VND)</h3>")
    parts.append("<table class='data'><thead><tr>"
                 "<th rowspan='2'>Channel</th>"
                 "<th colspan='4'>Disbursement</th>"
                 "<th colspan='4'>TOI</th>"
                 "<th colspan='4'>Provision (expense)</th>"
                 "<th colspan='4'>Profit Before Overhead</th>"
                 "</tr><tr>" +
                 ("<th>Actual</th><th>Budget</th><th>Δ vs BG</th><th>%mix</th>" * 4) +
                 "</tr></thead><tbody>")

    invert = [False, False, True, False]  # provision: +Δ = overshoot (bad)

    def row_channel(label, a_data, b_data, css_class=""):
        cells = [f"<td class='label {css_class}'>{label}</td>"]
        for i, m in enumerate(METRICS_4):
            a_val = a_data.get(m)
            b_val = b_data.get(m)
            d = (a_val - b_val) if (a_val is not None and b_val is not None) else None
            pct_d = (d/b_val*100) if (d is not None and b_val) else None
            mix = (a_val/(ov_a[m] or 1)*100) if (a_val is not None and ov_a[m]) else None
            cells.append(f"<td>{fmt(a_val)}</td><td>{fmt(b_val)}</td>")
            d_str = f"{fmt(d)} ({pct_d:+.0f}%)" if d is not None else "-"
            is_prov = invert[i]
            d_class = cell_class(d, neg_bad=False) if not is_prov else (("neg" if (d and d > 0) else "") if d is not None else "")
            cells.append(f"<td class='{d_class}'>{d_str}</td>")
            cells.append(f"<td class='accent'>{(f'{mix:.0f}%' if mix is not None else '-')}</td>")
        return "<tr class='" + css_class + "'>" + "".join(cells) + "</tr>"

    CHANNEL_LABELS = {"CAKE": "CAKE", "VDS": "VDS (incl. VDS-PR)", "ZLP": "ZLP (ZaloPay)", "MWG": "MWG"}
    for ch in GROUP_CHANNELS:
        parts.append(row_channel(CHANNEL_LABELS[ch], grp_a[ch], grp_b[ch]))
    parts.append(row_channel("Others", oth_a, oth_b))
    parts.append(row_channel("OVERALL Lending", ov_a, ov_b, css_class="total-row"))
    parts.append("</tbody></table>")

    # FY Budget context — annualized pace
    parts.append("<h4>FY 2026 Budget context (annualized pace check)</h4>")
    parts.append("<table class='data'><thead><tr><th>Channel</th><th>Disb FY Budget</th><th>Disb Actual annual. (×3)</th><th>FY pace</th>"
                 "<th>PBO FY Budget</th><th>PBO Actual annual.</th><th>FY pace</th></tr></thead><tbody>")
    pace_rows = [(CHANNEL_LABELS[ch], grp_a[ch], grp_bf[ch]) for ch in GROUP_CHANNELS]
    pace_rows.append(("Others", oth_a, oth_bf))
    pace_rows.append(("OVERALL", ov_a, ov_bf))
    for label, a_d, bf_d in pace_rows:
        da, pa = a_d.get("disb"), a_d.get("pbo")
        dbf, pbf = bf_d.get("disb"), bf_d.get("pbo")
        ann_d = (da * 3) if da is not None else None
        ann_p = (pa * 3) if pa is not None else None
        pace_d = (ann_d/dbf*100) if (ann_d is not None and dbf) else None
        pace_p = (ann_p/pbf*100) if (ann_p is not None and pbf) else None
        css = " total-row" if label == "OVERALL" else ""
        parts.append(f"<tr class='{css}'><td class='label'>{label}</td>"
                     f"<td>{fmt(dbf)}</td><td>{fmt(ann_d)}</td>"
                     f"<td class='{cell_class(pace_d-100 if pace_d else None, neg_bad=False)}'>{(f'{pace_d:.0f}%' if pace_d else '-')}</td>"
                     f"<td>{fmt(pbf)}</td><td>{fmt(ann_p)}</td>"
                     f"<td class='{cell_class(pace_p-100 if pace_p else None, neg_bad=False)}'>{(f'{pace_p:.0f}%' if pace_p else '-')}</td>"
                     f"</tr>")
    parts.append("</tbody></table>")
    parts.append("<p class='legend'>%mix = share of overall Lending. FY pace = (Actual YTD ×3) / Budget FY × 100. ZLP/MWG/Others budget chưa map riêng → pace '-'.</p>")

    # legacy vars for downstream sections (4.4 insights, gap attribution)
    cake_disb_a, vds_disb_a = grp_a["CAKE"]["disb"], grp_a["VDS"]["disb"]
    cake_disb_b, vds_disb_b = grp_b["CAKE"]["disb"], grp_b["VDS"]["disb"]
    oth_disb_a, oth_disb_b = oth_a["disb"], oth_b["disb"]
    ov_disb_a, ov_disb_b = ov_a["disb"], ov_b["disb"]

    parts.append("<h3>4.2 — Per-channel detail</h3>")

    # Per-channel breakdown table
    for ch in GROUP_CHANNELS:
        has_budget = ch in BUDGET_CHANNEL_ROWS
        parts.append(f"<h3>Channel: <code>{ch}</code>" + (" (with Budget comparison)" if has_budget else " (no Budget mapping yet)") + "</h3>")
        parts.append("<table class='data'><thead><tr>"
                     "<th>Product</th>"
                     "<th>Disb 25 FY</th><th>Disb 26 YTD</th>" + ("<th>Disb 26 BG YTD</th><th>Δ vs BG</th>" if has_budget else "") +
                     "<th>TOI 25 FY</th><th>TOI 26 YTD</th>" + ("<th>TOI 26 BG YTD</th><th>Δ vs BG</th>" if has_budget else "") +
                     "<th>PBO 25 FY</th><th>PBO 26 YTD</th>" + ("<th>PBO 26 BG YTD</th><th>Δ vs BG</th>" if has_budget else "") +
                     "<th>Yield 26</th><th>PBO mgn 26</th>"
                     "</tr></thead><tbody>")
        for product in ["CL", "OD", "Payday"]:
            d25 = deep[ch]["2025 FY"].get(product, {})
            d26 = deep[ch]["2026 YTD"].get(product, {})
            dB = deep[ch].get("2026 Budget YTD", {}).get(product, {})
            disb25, toi25, pbo25 = d25.get("disb"), d25.get("toi"), d25.get("pbo")
            disb26, toi26, pbo26 = d26.get("disb"), d26.get("toi"), d26.get("pbo")
            disbB, toiB, pboB = dB.get("disb"), dB.get("toi"), dB.get("pbo")
            if not any([disb25, disb26, disbB]):
                continue
            def delta_cell(actual, budget):
                if actual is None or budget is None: return "<td>-</td><td>-</td>"
                d = actual - budget
                pct = d/budget*100 if budget else 0
                return f"<td>{fmt(budget)}</td><td class='{cell_class(d, neg_bad=False)}'>{fmt(d)} ({pct:+.0f}%)</td>"
            yld26 = (toi26/disb26*100) if (disb26 and toi26 is not None) else None
            mgn26 = (pbo26/disb26*100) if (disb26 and pbo26 is not None) else None
            row_html = f"<tr><td class='label'>{product}</td>"
            row_html += f"<td>{fmt(disb25)}</td><td>{fmt(disb26)}</td>"
            if has_budget: row_html += delta_cell(disb26, disbB)
            row_html += f"<td>{fmt(toi25)}</td><td>{fmt(toi26)}</td>"
            if has_budget: row_html += delta_cell(toi26, toiB)
            row_html += f"<td class='{cell_class(pbo25)}'>{fmt(pbo25)}</td><td class='{cell_class(pbo26)}'>{fmt(pbo26)}</td>"
            if has_budget: row_html += delta_cell(pbo26, pboB)
            row_html += f"<td>{(f'{yld26:.1f}%' if yld26 is not None else '-')}</td>"
            row_html += f"<td class='{cell_class(mgn26)}'>{(f'{mgn26:.1f}%' if mgn26 is not None else '-')}</td>"
            row_html += "</tr>"
            parts.append(row_html)
        # Channel total
        def sum_metric(period, metric):
            return sum(v for p in deep[ch][period] for v in [deep[ch][period][p].get(metric)] if v is not None)
        tot_d25, tot_t25, tot_p25 = sum_metric("2025 FY","disb"), sum_metric("2025 FY","toi"), sum_metric("2025 FY","pbo")
        tot_d26, tot_t26, tot_p26 = sum_metric("2026 YTD","disb"), sum_metric("2026 YTD","toi"), sum_metric("2026 YTD","pbo")
        yld26 = tot_t26/tot_d26*100 if tot_d26 else None
        mgn26 = tot_p26/tot_d26*100 if tot_d26 else None
        row_html = f"<tr class='total-row'><td class='label'>Total {ch}</td>"
        row_html += f"<td>{fmt(tot_d25)}</td><td>{fmt(tot_d26)}</td>"
        if has_budget:
            tot_dB = sum_metric("2026 Budget YTD","disb")
            d = tot_d26 - tot_dB
            pct = d/tot_dB*100 if tot_dB else 0
            row_html += f"<td>{fmt(tot_dB)}</td><td class='{cell_class(d, neg_bad=False)}'>{fmt(d)} ({pct:+.0f}%)</td>"
        row_html += f"<td>{fmt(tot_t25)}</td><td>{fmt(tot_t26)}</td>"
        if has_budget:
            tot_tB = sum_metric("2026 Budget YTD","toi")
            d = tot_t26 - tot_tB
            pct = d/tot_tB*100 if tot_tB else 0
            row_html += f"<td>{fmt(tot_tB)}</td><td class='{cell_class(d, neg_bad=False)}'>{fmt(d)} ({pct:+.0f}%)</td>"
        row_html += f"<td>{fmt(tot_p25)}</td><td>{fmt(tot_p26)}</td>"
        if has_budget:
            tot_pB = sum_metric("2026 Budget YTD","pbo")
            d = tot_p26 - tot_pB
            pct = d/tot_pB*100 if tot_pB else 0
            row_html += f"<td>{fmt(tot_pB)}</td><td class='{cell_class(d, neg_bad=False)}'>{fmt(d)} ({pct:+.0f}%)</td>"
        row_html += f"<td>{(f'{yld26:.1f}%' if yld26 else '-')}</td><td>{(f'{mgn26:.1f}%' if mgn26 else '-')}</td></tr>"
        parts.append(row_html)
        parts.append("</tbody></table>")

    # VDS Channel vs CAKE comparison (VDS-PR đã merge vào VDS)
    parts.append("<h3>4.3 — VDS Channel vs CAKE — Yield/Margin profile</h3>")
    vds_family = {p: 0 for p in ["disb_25", "disb_26", "toi_25", "toi_26", "pbo_25", "pbo_26"]}
    cake = {p: 0 for p in ["disb_25", "disb_26", "toi_25", "toi_26", "pbo_25", "pbo_26"]}
    for ch_key, target in [("VDS", vds_family), ("CAKE", cake)]:
        for product in deep[ch_key]["2025 FY"]:
            d25 = deep[ch_key]["2025 FY"][product]
            d26 = deep[ch_key]["2026 YTD"][product]
            for m in ["disb", "toi", "pbo"]:
                if d25.get(m) is not None: target[f"{m}_25"] += d25[m]
                if d26.get(m) is not None: target[f"{m}_26"] += d26[m]

    parts.append("<table class='data'><thead><tr><th>Metric</th><th>VDS Channel</th><th>CAKE</th><th>Δ (CAKE − VDS)</th></tr></thead><tbody>")
    rows = [
        ("Disbursement 2025 FY", vds_family["disb_25"], cake["disb_25"]),
        (f"Disbursement 2026 {YTD_LABEL}", vds_family["disb_26"], cake["disb_26"]),
        (f"TOI 2026 {YTD_LABEL}", vds_family["toi_26"], cake["toi_26"]),
        (f"PBO 2026 {YTD_LABEL}", vds_family["pbo_26"], cake["pbo_26"]),
    ]
    for label, vds, cak in rows:
        diff = cak - vds
        parts.append(f"<tr><td class='label'>{label}</td><td>{fmt(vds)}</td><td>{fmt(cak)}</td>"
                     f"<td class='{cell_class(diff, neg_bad=False)}'>{fmt(diff)}</td></tr>")
    # Yield + margin comparison
    yld_v_25 = vds_family["toi_25"]/vds_family["disb_25"]*100
    yld_v_26 = vds_family["toi_26"]/vds_family["disb_26"]*100
    yld_c_25 = cake["toi_25"]/cake["disb_25"]*100
    yld_c_26 = cake["toi_26"]/cake["disb_26"]*100
    mgn_v_25 = vds_family["pbo_25"]/vds_family["disb_25"]*100
    mgn_v_26 = vds_family["pbo_26"]/vds_family["disb_26"]*100
    mgn_c_25 = cake["pbo_25"]/cake["disb_25"]*100
    mgn_c_26 = cake["pbo_26"]/cake["disb_26"]*100
    parts.append(f"<tr><td class='label'>Yield (TOI/Disb) 2025</td><td>{yld_v_25:.1f}%</td><td>{yld_c_25:.1f}%</td><td>{yld_c_25-yld_v_25:+.1f} pp</td></tr>")
    parts.append(f"<tr><td class='label'>Yield 2026 YTD</td><td>{yld_v_26:.1f}%</td><td>{yld_c_26:.1f}%</td><td>{yld_c_26-yld_v_26:+.1f} pp</td></tr>")
    parts.append(f"<tr><td class='label'>PBO margin 2025</td><td>{mgn_v_25:.1f}%</td><td>{mgn_c_25:.1f}%</td><td class='pos-neg'>{mgn_c_25-mgn_v_25:+.1f} pp</td></tr>")
    parts.append(f"<tr><td class='label'>PBO margin 2026 YTD</td><td>{mgn_v_26:.1f}%</td><td>{mgn_c_26:.1f}%</td><td class='pos-neg'>{mgn_c_26-mgn_v_26:+.1f} pp</td></tr>")
    parts.append("</tbody></table>")

    # Channel Risk Profile (NPL CAKE + NPL CIC + LG2)
    parts.append("<h3>4.3.5 — Channel Risk Profile (NPL CAKE + NPL CIC + Loan group 2)</h3>")
    parts.append("<p class='note'>Risk state at end-of-period: 2025 FY = Dec 25, 2026 YTD = Apr 26. <strong>NPL CAKE</strong> = định nghĩa nội bộ Cake. <strong>NPL CIC</strong> = định nghĩa CIC (thường cao hơn). LG2 = % balance overdue 10-90 days (leading indicator).</p>")
    parts.append("<table class='data'><thead><tr><th>Channel</th><th>Product</th>"
                 "<th>NPL CAKE 25</th><th>NPL CAKE 26</th>"
                 "<th>NPL CIC 25</th><th>NPL CIC 26</th>"
                 "<th>LG2 25</th><th>LG2 26</th>"
                 "<th>Balance Dec25 (Tỷ)</th><th>Balance Apr26 (Tỷ)</th></tr></thead><tbody>")
    def _pct2(v):
        return f"{v*100:.2f}%" if v is not None else "-"
    for ch in GROUP_CHANNELS:
        for product in ["CL", "OD", "Payday"]:
            if product not in deep[ch]["2025 FY"]:
                continue
            d25 = deep[ch]["2025 FY"][product]
            d26 = deep[ch]["2026 YTD"][product]
            nc25 = d25.get("npl_cake"); nc26 = d26.get("npl_cake")
            ni25 = d25.get("npl_cic"); ni26 = d26.get("npl_cic")
            lg2_25 = d25.get("lg2"); lg2_26 = d26.get("lg2")
            bal25 = d25.get("balance"); bal26 = d26.get("balance")
            cls_nc = "neg" if (nc25 is not None and nc26 is not None and nc26 > nc25) else ""
            cls_ni = "neg" if (ni25 is not None and ni26 is not None and ni26 > ni25) else ""
            cls_lg = "neg" if (lg2_25 is not None and lg2_26 is not None and lg2_26 > lg2_25) else ""
            parts.append(f"<tr><td class='label'>{ch}</td><td class='label'>{product}</td>"
                         f"<td>{_pct2(nc25)}</td><td class='{cls_nc}'>{_pct2(nc26)}</td>"
                         f"<td>{_pct2(ni25)}</td><td class='{cls_ni}'>{_pct2(ni26)}</td>"
                         f"<td>{_pct2(lg2_25)}</td><td class='{cls_lg}'>{_pct2(lg2_26)}</td>"
                         f"<td>{fmt(bal25)}</td><td>{fmt(bal26)}</td></tr>")
    parts.append("</tbody></table>")

    # Net Yield (TOI − Risk) comparison + Partner Sharing context
    parts.append("<h3>4.3.6 — Net Yield (TOI − Risk) by Channel + Partner Sharing</h3>")
    parts.append("<p class='note'><strong>Tại sao cần Net Yield:</strong> raw TOI comparison giữa các channel bị misleading vì <strong>mỗi partner có cấu trúc sharing khác nhau</strong>. Net Yield = TOI − Provision (Risk) là metric so sánh fair hơn. Lưu ý sharing params dưới đây chỉ ảnh hưởng economics thực tế Cake giữ lại.</p>")
    sharing = load_channel_sharing_params()
    parts.append("<table class='data'><thead><tr><th>Channel</th><th>Product</th>"
                 "<th>TOI 2025 FY</th><th>Risk (Provision)</th><th>Net Yield (TOI−Risk)</th>"
                 "<th>Net Yield /Disb</th><th>Net Yield /ANR</th>"
                 "<th>Partner Sharing</th></tr></thead><tbody>")
    SHARING_CH_MAP = {"ZLP": "ZALOPAY"}  # GROUP_CHANNELS name → sharing-file channel name
    for ch in GROUP_CHANNELS:
        for product in ["CL", "OD", "Payday"]:
            if product not in deep[ch]["2025 FY"]:
                continue
            d25 = deep[ch]["2025 FY"][product]
            toi = d25.get("toi")
            prov = d25.get("prov")
            disb = d25.get("disb")
            bal = d25.get("balance")
            if toi is None or prov is None:
                continue
            net_yield = toi - prov
            ny_disb = (net_yield / disb * 100) if disb else None
            ny_anr = (net_yield / bal * 100) if bal else None
            # Sharing param lookup (CL/PD only; map Payday→PD, ZLP→ZALOPAY)
            shr_product = "PD" if product == "Payday" else product
            shr = sharing.get(shr_product, {}).get(SHARING_CH_MAP.get(ch, ch), {})
            shr_nonzero = {k.replace("sharing_", "").replace("_", " "): v for k, v in shr.items() if v}
            if shr_nonzero:
                shr_str = "; ".join(f"{k}={v:.1%}" if v < 1 else f"{k}={v}" for k, v in shr_nonzero.items())
            else:
                shr_str = "<span class='na'>no sharing (Cake 100%)</span>"
            parts.append(f"<tr><td class='label'>{ch}</td><td class='label'>{product}</td>"
                         f"<td>{fmt(toi)}</td><td class='neg'>{fmt(prov)}</td>"
                         f"<td class='accent'>{fmt(net_yield)}</td>"
                         f"<td>{(f'{ny_disb:.1f}%' if ny_disb is not None else '-')}</td>"
                         f"<td>{(f'{ny_anr:.1f}%' if ny_anr is not None else '-')}</td>"
                         f"<td><small>{shr_str}</small></td></tr>")
    parts.append("</tbody></table>")
    parts.append("<p class='legend'><strong>Đọc bảng:</strong> Net Yield /Disb cao = channel sinh lời tốt sau khi trừ rủi ro. Nhưng nếu partner sharing cao (vd VDS share 50% TOI+Risk từ DPD180+), thì Net Yield Cake thực giữ lại còn thấp hơn con số hiển thị. Sharing chi tiết xem JSON section 7.</p>")

    # Observations
    parts.append("<h3>4.4 — Insights</h3>")
    # 3-way disb gap breakdown
    cake_disb_gap = cake_disb_a - cake_disb_b
    vds_disb_gap = vds_disb_a - vds_disb_b
    oth_disb_gap = oth_disb_a - oth_disb_b
    total_gap = ov_disb_a - ov_disb_b
    parts.append(f"<p class='note'><strong>Disbursement gap attribution (Tỷ VND, {YTD_LABEL}):</strong><br>"
                 f"Total miss = <strong>{total_gap:+,.0f}</strong> ({total_gap/ov_disb_b*100:+.0f}%)<br>"
                 f"&nbsp;&nbsp;• CAKE: <strong>{cake_disb_gap:+,.0f}</strong> ({cake_disb_gap/total_gap*100:.0f}% of gap)<br>"
                 f"&nbsp;&nbsp;• VDS: <strong>{vds_disb_gap:+,.0f}</strong> ({vds_disb_gap/total_gap*100:.0f}% of gap)<br>"
                 f"&nbsp;&nbsp;• Others: <strong>{oth_disb_gap:+,.0f}</strong> ({oth_disb_gap/total_gap*100:.0f}% of gap)</p>")
    parts.append("<ol class='insights'>")
    parts.append(f"<li><strong>Volume gần như ngang nhau, nhưng profit khác hẳn.</strong> "
                 f"2025 FY Disbursement: VDS Channel {vds_family['disb_25']:,.0f} vs CAKE {cake['disb_25']:,.0f}. "
                 f"PBO: VDS {vds_family['pbo_25']:,.0f} vs CAKE {cake['pbo_25']:,.0f} → CAKE cao hơn {cake['pbo_25']-vds_family['pbo_25']:,.0f} Tỷ "
                 f"({((cake['pbo_25']-vds_family['pbo_25'])/vds_family['pbo_25']*100):+.0f}%).</li>")
    parts.append(f"<li><strong>CAKE margin gấp đôi VDS</strong> ({mgn_c_26:.1f}% vs {mgn_v_26:.1f}% PBO/Disb 2026 YTD). "
                 f"VDS yield cao hơn ({yld_v_26:.1f}% vs {yld_c_26:.1f}%) nhưng partner acquisition cost + commission ăn margin.</li>")
    parts.append(f"<li><strong>CAKE Payday đang explode:</strong> 2026 {YTD_LABEL} Disbursement {cake.get('disb_26', 0):,.0f} (riêng PD: "
                 f"{deep['CAKE']['2026 YTD']['Payday']['disb']:,.0f}) — chỉ {ACTUAL_2026_COMPLETE_MONTHS} tháng đã vượt PD CAKE cả năm 2025 ({deep['CAKE']['2025 FY']['Payday']['disb']:,.0f}). "
                 f"Annualized YoY +{(deep['CAKE']['2026 YTD']['Payday']['disb']*ANNUALIZE - deep['CAKE']['2025 FY']['Payday']['disb'])/deep['CAKE']['2025 FY']['Payday']['disb']*100:.0f}%.</li>")
    parts.append(f"<li><strong>CAKE OD lỗ kéo dài:</strong> PBO 2025 FY {deep['CAKE']['2025 FY']['OD']['pbo']:+.1f}, "
                 f"2026 YTD {deep['CAKE']['2026 YTD']['OD']['pbo']:+.1f}. "
                 f"Volume 2026 annualized = {deep['CAKE']['2026 YTD']['OD']['disb']*ANNUALIZE:.0f} vs 2025 {deep['CAKE']['2025 FY']['OD']['disb']:.0f} → {(deep['CAKE']['2026 YTD']['OD']['disb']*ANNUALIZE - deep['CAKE']['2025 FY']['OD']['disb'])/deep['CAKE']['2025 FY']['OD']['disb']*100:+.0f}% YoY. "
                 f"Đang cố ý down-scale hay mất market?</li>")
    parts.append(f"<li><strong>CAKE CL margin compression:</strong> "
                 f"PBO/Disb CL: 2025 = {deep['CAKE']['2025 FY']['CL']['pbo']/deep['CAKE']['2025 FY']['CL']['disb']*100:.1f}%, "
                 f"2026 YTD = {deep['CAKE']['2026 YTD']['CL']['pbo']/deep['CAKE']['2026 YTD']['CL']['disb']*100:.1f}%. "
                 f"Provision tăng / mix shift sang segment yield thấp / hay one-off recovery 2025?</li>")
    parts.append("</ol>")

    return "\n".join(parts)


# ========== BY CHANNEL (2025 Actual FY × Product matrix) ==========
def load_channel_matrix_2025():
    """Returns {channel: {product: {disb, toi, pbo}}} for FY 2025."""
    wb = openpyxl.load_workbook(RAW/"P&L_2025_final.xlsx", data_only=True)
    ws = wb["Lending"]
    matrix = {}
    for product, bases in ACTUAL_2025_PRODUCT_CHANNEL_BASES.items():
        for base in bases:
            channel = CHANNEL_NAMES_2025[base]
            if channel not in matrix:
                matrix[channel] = {}
            if product not in matrix[channel]:
                matrix[channel][product] = {}
            for metric in ["disb", "toi", "pbo"]:
                row = base + PRODUCT_CHANNEL_OFFSETS[metric]
                monthly = [num(ws.cell(row, c).value, 1e9) for c in range(16, 28)]
                matrix[channel][product][metric] = safe_sum(monthly)
    wb.close()
    return matrix


# ========== HTML rendering ==========
def fmt(v, decimals=1):
    if v is None: return "<span class='na'>-</span>"
    return f"{v:,.{decimals}f}"


def cell_class(v, neg_bad=True):
    if v is None: return ""
    if neg_bad and v < 0: return "neg"
    if v < 0: return "pos-neg"
    return ""


def render_overall_table(overall, periods=None, show_delta=True):
    """Render overall P&L tables. periods = subset of ['2025 Actual','2026 Actual','2026 Budget']."""
    if periods is None:
        periods = ["2025 Actual", "2026 Actual", "2026 Budget"]
    metrics = [("disb", "Disbursement"), ("toi", "TOI"),
               ("prov", "Provision expense"),
               ("pbo", "Profit before overhead"), ("pbt", "PBT")]
    html_parts = []
    for metric_key, metric_label in metrics:
        html_parts.append(f"<h3>{metric_label}</h3>")
        html_parts.append("<table class='data'>")
        html_parts.append("<thead><tr><th>Period</th>")
        for m in MONTHS:
            html_parts.append(f"<th>{m}</th>")
        html_parts.append(f"<th class='accent'>{YTD_LABEL}</th><th class='accent'>FY</th></tr></thead><tbody>")

        ytd_n = ACTUAL_2026_COMPLETE_MONTHS
        for period in periods:
            vals = overall[period][metric_key]
            ytd = safe_sum(vals[:ytd_n])
            fy = safe_sum(vals)
            html_parts.append(f"<tr><td class='label'>{period}</td>")
            for v in vals:
                html_parts.append(f"<td class='{cell_class(v)}'>{fmt(v)}</td>")
            html_parts.append(f"<td class='accent {cell_class(ytd)}'>{fmt(ytd)}</td>")
            html_parts.append(f"<td class='accent {cell_class(fy)}'>{fmt(fy)}</td></tr>")
        # Δ vs Budget YTD — only if both 2026 Actual + Budget in periods
        if show_delta and "2026 Actual" in periods and "2026 Budget" in periods:
            a_ytd = safe_sum(overall["2026 Actual"][metric_key][:ytd_n])
            b_ytd = safe_sum(overall["2026 Budget"][metric_key][:ytd_n])
            if a_ytd is not None and b_ytd:
                d = a_ytd - b_ytd
                pct = d / b_ytd * 100
                html_parts.append(f"<tr class='delta'><td class='label'>Δ Actual vs Budget {YTD_LABEL}</td>"
                                  f"<td colspan='12'></td>"
                                  f"<td class='accent {cell_class(d)}'>{fmt(d)} ({pct:+.0f}%)</td><td></td></tr>")
        html_parts.append("</tbody></table>")
    return "\n".join(html_parts)


def render_by_product(by_product):
    metrics = [("disb", "Disbursement"), ("toi", "TOI"), ("pbo", "Profit before overhead")]
    parts = []
    for metric_key, metric_label in metrics:
        parts.append(f"<h3>{metric_label} — by Product</h3>")
        parts.append("<table class='data'>")
        parts.append("<thead><tr><th>Product</th><th>2025 Actual FY</th>"
                     f"<th>2026 Actual {YTD_LABEL}</th><th>2026 Budget {YTD_LABEL}</th>"
                     "<th>Δ vs BG YTD</th><th>YoY vs 2025 YTD</th>"
                     "<th>2026 Budget FY</th><th>% FY done</th></tr></thead><tbody>")

        for product in ["CashLoan", "Overdraft", "Payday", "Paylater"]:
            metric_data = by_product[product][metric_key]
            a25_fy = safe_sum(metric_data.get("2025 Actual", [None]*12))
            a25_ytd = safe_sum(metric_data.get("2025 Actual", [None]*12)[:ACTUAL_2026_COMPLETE_MONTHS])
            a26_ytd = safe_sum(metric_data.get("2026 Actual", [None]*12)[:ACTUAL_2026_COMPLETE_MONTHS])
            b26_ytd = safe_sum(metric_data.get("2026 Budget", [None]*12)[:ACTUAL_2026_COMPLETE_MONTHS])
            b26_fy = safe_sum(metric_data.get("2026 Budget", [None]*12))
            delta = (a26_ytd - b26_ytd) if (a26_ytd is not None and b26_ytd is not None) else None
            delta_pct = (delta/b26_ytd*100) if (delta is not None and b26_ytd) else None
            yoy = (a26_ytd - a25_ytd) if (a26_ytd is not None and a25_ytd is not None) else None
            yoy_pct = (yoy/a25_ytd*100) if (yoy is not None and a25_ytd) else None
            pace = (a26_ytd/b26_fy*100) if (a26_ytd is not None and b26_fy) else None
            parts.append(f"<tr><td class='label'>{product}</td>")
            parts.append(f"<td>{fmt(a25_fy)}</td>")
            parts.append(f"<td>{fmt(a26_ytd)}</td>")
            parts.append(f"<td>{fmt(b26_ytd)}</td>")
            d_str = f"{fmt(delta)} ({delta_pct:+.0f}%)" if delta is not None else "<span class='na'>-</span>"
            parts.append(f"<td class='{cell_class(delta)}'>{d_str}</td>")
            y_str = f"{fmt(yoy)} ({yoy_pct:+.0f}%)" if yoy is not None else "<span class='na'>-</span>"
            parts.append(f"<td class='{cell_class(yoy)}'>{y_str}</td>")
            parts.append(f"<td>{fmt(b26_fy)}</td>")
            parts.append(f"<td>{fmt(pace)}%</td></tr>" if pace is not None else "<td>-</td></tr>")
        parts.append("</tbody></table>")
    return "\n".join(parts)


# Raw channel → group mapping (theo Đạm: CAKE / VDS / ZLP / MWG / Others)
CHANNEL_GROUP_MAP = {
    "CAKE": "CAKE",
    "VDS": "VDS", "VDS-PR": "VDS",
    "ZALOPAY": "ZLP", "ZLP": "ZLP",
    "MWG": "MWG",
    "BEG": "Others", "NGS": "Others", "VNPAY": "Others",
    "VNPOST": "Others", "VNPOST-PR": "Others", "MISA": "Others",
}
GROUP_DISPLAY_ORDER = ["CAKE", "VDS", "ZLP", "MWG", "Others"]


def render_by_channel(matrix):
    """2025 FY channel × product matrix — grouped CAKE/VDS/ZLP/MWG/Others (VDS-PR merged vào VDS)."""
    metrics = [("disb", "Disbursement"), ("toi", "TOI"), ("pbo", "Profit before overhead")]
    products_order = ["CashLoan", "Overdraft", "Payday"]
    parts = []
    parts.append("<p class='note'>Channel breakdown <strong>2025 Actual FY</strong>, grouped theo Đạm: CAKE / VDS (incl VDS-PR) / ZLP / MWG / Others. "
                 "Raw channels gộp: VDS-PR→VDS, ZALOPAY→ZLP; BEG/NGS/VNPAY/VNPOST/VNPOST-PR/MISA→Others.</p>")
    for metric_key, metric_label in metrics:
        parts.append(f"<h3>{metric_label} — 2025 Actual FY (Tỷ VND)</h3>")
        parts.append("<table class='data'><thead><tr><th>Channel</th>")
        for p in products_order:
            parts.append(f"<th>{p}</th>")
        parts.append("<th class='accent'>Total</th></tr></thead><tbody>")
        # Aggregate raw channels into groups
        grouped = {g: {p: 0.0 for p in products_order} for g in GROUP_DISPLAY_ORDER}
        for ch, products_dict in matrix.items():
            group = CHANNEL_GROUP_MAP.get(ch, "Others")
            for p in products_order:
                v = products_dict.get(p, {}).get(metric_key)
                if v is not None:
                    grouped[group][p] += v
        col_totals = {p: 0.0 for p in products_order}
        grand = 0.0
        for group in GROUP_DISPLAY_ORDER:
            parts.append(f"<tr><td class='label'>{group}</td>")
            row_total = 0.0
            for p in products_order:
                v = grouped[group][p]
                row_total += v
                col_totals[p] += v
                parts.append(f"<td class='{cell_class(v)}'>{fmt(v) if v else '-'}</td>")
            grand += row_total
            parts.append(f"<td class='accent'>{fmt(row_total) if row_total else '-'}</td></tr>")
        parts.append(f"<tr class='total-row'><td class='label'>Total</td>")
        for p in products_order:
            parts.append(f"<td class='accent'>{fmt(col_totals[p])}</td>")
        parts.append(f"<td class='accent'>{fmt(grand)}</td></tr>")
        parts.append("</tbody></table>")
    return "\n".join(parts)


# ========== UNIT ECONOMICS (Section 5) ==========
# Q425 Summary file: Lending aggregate (UPL) + per-product (CL/PD/PL/OD/CC) per-loan UE
UE_SUM_PATH = RAW.parent / "Unit Economic" / "Q425" / "Unit Economics Q42025 Lending.xlsx"
Q126_PATH = RAW.parent / "Unit Economic" / "Q126"

# UPL aggregate Lending: SUM sheet rows R10-R17, Q4.25=col 4, Q3.25=col 6
# Note: UE SUM file has single "NPL" row (R16) — không tách CAKE/CIC → map vào npl_cake, npl_cic = None
UE_SUM_ROWS = {
    "disb":      10,  # Disburse/Spending
    "balance":   11,  # Balance ≈ ANR proxy
    "toi":       12,
    "provision": 13,
    "opex":      14,
    "pbt":       15,
    "npl_cake":  16,  # single NPL in SUM file — treated as CAKE definition
    "lg2":       17,
}

# CL per-loan UE in CL sheet: rows R3-R16 = Q4.25 block, R21-R34 = Q3.25 block
# Channel cols: CL all=3, CAKE=5, VDS=7, MWG=9, VNPAY=11, ZLP=13
CL_SHEET_METRIC_ROWS_Q4 = {
    "anr":           3,
    "toi":           4,
    "nii":           5,
    "nfi":           6,
    "provision":     8,
    "opex":          9,
    "acquisition":   10,
    "uw_disbursal":  11,
    "collection":    12,
    "aftersale":     13,
    "support_unit":  14,
    "pbt":           15,
    "pbsc":          16,
}
CL_SHEET_METRIC_ROWS_Q3 = {k: v + 18 for k, v in CL_SHEET_METRIC_ROWS_Q4.items()}
CL_CHANNEL_COLS = {  # channel → (value_col, %ANR_col)
    "CL all": (3, 4),
    "CAKE":   (5, 6),
    "VDS":    (7, 8),
    "MWG":    (9, 10),
    "VNPAY":  (11, 12),
    "ZLP":    (13, 14),
}

# Q1.26 CL channel inputs from CashLoan_Q1_26.xlsx DATA sheet
# Row 44 header: cols 7=Approval, 8=Signed, 9=IR, 10=COF, 11=AVG Ticket, 12=AVG Tenor, 13=Insurance rate, 14=Insurance pen
# Rows 45-52: BE, Cake, MISA, MWG, VNPOST, VDS, VNPAY, ZALOPAY; R53=All
Q126_CL_DATA_ROWS = {
    "BE":      45,
    "CAKE":    46,
    "MISA":    47,
    "MWG":     48,
    "VNPOST":  49,
    "VDS":     50,
    "VNPAY":   51,
    "ZALOPAY": 52,
    "All":     53,
}
Q126_CL_DATA_COLS = {
    "approval_rate":  7,
    "signed_rate":    8,
    "ir":             9,
    "cof":            10,
    "avg_ticket":     11,
    "avg_tenor":      12,
    "insurance_rate": 13,
    "insurance_pen":  14,
}


def load_ue_q3q4_25():
    """Returns {quarter: {metric: value}} for Lending aggregate UE."""
    wb = openpyxl.load_workbook(UE_SUM_PATH, data_only=True)
    ws = wb["SUM"]
    out = {"Q3.25": {}, "Q4.25": {}}
    for metric, row in UE_SUM_ROWS.items():
        q4 = ws.cell(row, 4).value  # Q4.25
        q3 = ws.cell(row, 6).value  # Q3.25
        # Provision/OPEX stored as negative (expense); normalize to abs for ratio display
        if metric in {"provision", "opex"}:
            q4 = abs(q4) if q4 is not None else None
            q3 = abs(q3) if q3 is not None else None
        out["Q4.25"][metric] = q4
        out["Q3.25"][metric] = q3
    # UE SUM file không tách NPL CIC → mark None cho cả 2 quý
    out["Q4.25"]["npl_cic"] = None
    out["Q3.25"]["npl_cic"] = None
    wb.close()
    return out


def load_cl_per_loan_ue():
    """Returns {channel: {quarter: {metric: (value, %ANR)}}} per-loan UE for CashLoan.
    Quarter = 'Q3.25' or 'Q4.25'. Source: Q425 SUM file CL sheet."""
    wb = openpyxl.load_workbook(UE_SUM_PATH, data_only=True)
    ws = wb["CL"]
    out = {}
    for channel, (val_col, pct_col) in CL_CHANNEL_COLS.items():
        out[channel] = {"Q4.25": {}, "Q3.25": {}}
        for metric, row_q4 in CL_SHEET_METRIC_ROWS_Q4.items():
            row_q3 = CL_SHEET_METRIC_ROWS_Q3[metric]
            v_q4 = ws.cell(row_q4, val_col).value
            p_q4 = ws.cell(row_q4, pct_col).value
            v_q3 = ws.cell(row_q3, val_col).value
            p_q3 = ws.cell(row_q3, pct_col).value
            # Normalize provision/opex/sub-opex to abs for ratio display consistency
            if metric in {"provision", "opex", "acquisition", "uw_disbursal", "collection", "aftersale", "support_unit"}:
                v_q4 = abs(v_q4) if v_q4 is not None else None
                v_q3 = abs(v_q3) if v_q3 is not None else None
                p_q4 = abs(p_q4) if p_q4 is not None else None
                p_q3 = abs(p_q3) if p_q3 is not None else None
            out[channel]["Q4.25"][metric] = (v_q4, p_q4)
            out[channel]["Q3.25"][metric] = (v_q3, p_q3)
    wb.close()
    return out


def load_by_product_quarterly():
    """Returns {product: {quarter: {metric: value}}}.
    Numeric metrics (Tỷ VND): disb, toi, provision, pbo. Risk metrics (%): npl, lg2.
    Quarter = 'Q3.25' | 'Q4.25' | 'Q1.26'.
    Source: P&L_2025_final.xlsx (channel-summed for product aggregate) + P&L_2026_Actual.xlsx product sheets."""
    PRODUCT_CHANNEL_ROWS_2025 = {
        "CL": [52, 91, 130, 169, 208, 247, 286, 325, 364, 403, 442],
        "PD": [520, 559, 598, 637, 675],
        "OD": [481],
        "PL": [],  # Paylater not split per channel in 2025 Lending sheet — Q1.26 only from 2026 file
    }
    OFFSET_2025 = {"disb": 0, "toi": 12, "provision": 23, "pbo": 33}
    # NPL CAKE +10, NPL CIC +11 (channel-level % rows); LG2 numerator (Overdue 11-90d) +9; Balance ENR +3
    OFFSET_2025_NPL_CAKE_PCT = 10
    OFFSET_2025_NPL_CIC_PCT = 11
    OFFSET_2025_LG2_BAL = 9
    OFFSET_2025_BALANCE_ENR = 3

    PRODUCT_2026 = {
        "CL":     {"sheet": "CL",       "disb": 6, "toi": 19, "provision": 30, "pbo": 40, "lg2": 16, "npl_cake": 17, "npl_cic": 18, "balance": 9},
        "PD":     {"sheet": "Payday",   "disb": 6, "toi": 19, "provision": 30, "pbo": 40, "lg2": 16, "npl_cake": 17, "npl_cic": 18, "balance": 9},
        "OD":     {"sheet": "OD",       "disb": 6, "toi": 20, "provision": 31, "pbo": 41, "lg2": 17, "npl_cake": 18, "npl_cic": 19, "balance": 9},
        "PL":     {"sheet": "Paylater", "disb": 8, "toi": 25, "provision": 42, "pbo": 52, "lg2": 20, "npl_cake": 22, "npl_cic": 21, "balance": 9},  # Spending=R8, Principle balance=R9
    }

    all_products = list(PRODUCT_CHANNEL_ROWS_2025.keys())
    out = {p: {"Q3.25": {}, "Q4.25": {}, "Q1.26": {}} for p in all_products}

    # 2025: aggregate channels per product per quarter (P&L flows + risk state metrics)
    wb25 = openpyxl.load_workbook(RAW/"P&L_2025_final.xlsx", data_only=True)
    ws25 = wb25["Lending"]
    # Q3.25 = cols 22-24 (Jul-Sep), Q4.25 = cols 25-27 (Oct-Dec). NPL/LG2 use last col only (state metric).
    Q3_COLS = [22, 23, 24]
    Q4_COLS = [25, 26, 27]
    for product, channel_bases in PRODUCT_CHANNEL_ROWS_2025.items():
        if not channel_bases:
            # Product not in 2025 Lending sheet (e.g., PL) — mark Q3/Q4 as None
            for quarter in ["Q3.25", "Q4.25"]:
                for metric in ["disb", "toi", "provision", "pbo", "balance", "npl_cake", "npl_cic", "lg2"]:
                    out[product][quarter][metric] = None
            continue
        for quarter, qcols in [("Q3.25", Q3_COLS), ("Q4.25", Q4_COLS)]:
            # Flow metrics (sum across months)
            for metric, off in OFFSET_2025.items():
                total = 0
                for base in channel_bases:
                    row = base + off
                    for c in qcols:
                        v = ws25.cell(row, c).value
                        if v is not None and not isinstance(v, str):
                            total += v
                total_bil = total / 1e9 if total != 0 else 0
                if metric == "provision":
                    total_bil = abs(total_bil)
                out[product][quarter][metric] = total_bil
            # NPL CAKE + CIC: balance-weighted average across channels (state at quarter end, last col)
            end_col = qcols[-1]
            total_overdue = 0       # sum overdue 11-90d (LG2 numerator) raw VND
            total_balance = 0       # sum balance ENR (denominator) raw VND
            total_npl_cake_bal = 0  # numerator for weighted NPL CAKE
            total_npl_cic_bal = 0   # numerator for weighted NPL CIC
            for base in channel_bases:
                bal = ws25.cell(base + OFFSET_2025_BALANCE_ENR, end_col).value
                npl_cake = ws25.cell(base + OFFSET_2025_NPL_CAKE_PCT, end_col).value
                npl_cic = ws25.cell(base + OFFSET_2025_NPL_CIC_PCT, end_col).value
                lg2_bal = ws25.cell(base + OFFSET_2025_LG2_BAL, end_col).value
                if bal is not None and not isinstance(bal, str):
                    total_balance += bal
                    if npl_cake is not None and not isinstance(npl_cake, str):
                        total_npl_cake_bal += npl_cake * bal
                    if npl_cic is not None and not isinstance(npl_cic, str):
                        total_npl_cic_bal += npl_cic * bal
                    if lg2_bal is not None and not isinstance(lg2_bal, str):
                        total_overdue += lg2_bal
            out[product][quarter]["balance"] = total_balance / 1e9 if total_balance else 0
            out[product][quarter]["npl_cake"] = (total_npl_cake_bal / total_balance) if total_balance else None
            out[product][quarter]["npl_cic"] = (total_npl_cic_bal / total_balance) if total_balance else None
            out[product][quarter]["lg2"] = (total_overdue / total_balance) if total_balance else None
    wb25.close()

    # 2026 Q1.26: from product sheets cols 6+7+8 (Jan+Feb+Mar). NPL/LG2 = Mar (col 8) state.
    wb26 = openpyxl.load_workbook(RAW/"P&L_2026_Actual.xlsx", data_only=True)
    Q1_COLS = [6, 7, 8]
    Q1_END_COL = 8  # Mar 2026 for state metrics
    for product, cfg in PRODUCT_2026.items():
        ws = wb26[cfg["sheet"]]
        # Flow metrics
        for metric in ["disb", "toi", "provision", "pbo"]:
            row = cfg[metric]
            total = 0
            for c in Q1_COLS:
                v = ws.cell(row, c).value
                if v is not None and not isinstance(v, str):
                    total += v
            total_bil = total / 1e9 if total != 0 else 0
            if metric == "provision":
                total_bil = abs(total_bil)
            out[product]["Q1.26"][metric] = total_bil
        # State metrics: NPL CAKE/CIC, LG2%, Balance ENR (end of Q1)
        def _q1state(row):
            v = ws.cell(row, Q1_END_COL).value
            return v if (v is not None and not isinstance(v, str)) else None
        balance_v = ws.cell(cfg["balance"], Q1_END_COL).value
        out[product]["Q1.26"]["balance"] = (balance_v / 1e9) if (balance_v and not isinstance(balance_v, str)) else None
        out[product]["Q1.26"]["npl_cake"] = _q1state(cfg["npl_cake"])
        out[product]["Q1.26"]["npl_cic"] = _q1state(cfg["npl_cic"])
        out[product]["Q1.26"]["lg2"] = _q1state(cfg["lg2"])
    wb26.close()
    return out


def load_q126_cl_channel_inputs():
    """Returns {channel: {input_metric: value}} from Q126 CashLoan DATA sheet."""
    wb = openpyxl.load_workbook(Q126_PATH / "CashLoan_Q1_26.xlsx", data_only=True)
    ws = wb["DATA"]
    out = {}
    for channel, row in Q126_CL_DATA_ROWS.items():
        out[channel] = {}
        for metric, col in Q126_CL_DATA_COLS.items():
            out[channel][metric] = ws.cell(row, col).value
    wb.close()
    return out


# Q1.26 Payday channel inputs from Payday_Q1_26.xlsx DATA sheet
# Row 40 header, R41-R44 channels, R45 All
Q126_PD_DATA_ROWS = {
    "VDS":     41,
    "CAKE":    42,
    "VNPAY":   43,
    "ZALOPAY": 44,
    "All":     45,
}
Q126_PD_DATA_COLS = {  # same column layout as CL
    "approval_rate":  7,
    "signed_rate":    8,
    "ir":             9,
    "cof":            10,
    "avg_ticket":     11,
    "avg_tenor":      12,
    "insurance_rate": 13,
    "insurance_pen":  14,
}


def load_q126_pd_channel_inputs():
    """Returns {channel: {input_metric: value}} from Q126 Payday DATA sheet."""
    wb = openpyxl.load_workbook(Q126_PATH / "Payday_Q1_26.xlsx", data_only=True)
    ws = wb["DATA"]
    out = {}
    for channel, row in Q126_PD_DATA_ROWS.items():
        out[channel] = {}
        for metric, col in Q126_PD_DATA_COLS.items():
            out[channel][metric] = ws.cell(row, col).value
    wb.close()
    return out


PAYLATER_CHANNEL_BASES = {  # row of channel header (e.g., R54 = "VDS paylater"). Top-level uses base 1 (R2 = base+1).
    "Total":      1,
    "VDS":        54,
    "VDS-EPASS":  107,
    "BE":         160,
    "VNPAY":      213,
    "FPT":        266,
    "MWG":        319,
}
PAYLATER_OFFSETS = {  # all offsets from base header row
    "accumulated_account": 2,
    "active_account":      3,
    "spending":            7,   # raw VND, flow
    "balance":             8,   # raw VND, state (Principle balance)
    "lg2":                 19,  # % Attention debts (often 0/empty in this file — fallback computed)
    "npl_cic":             20,  # %
    "npl_cake":            21,  # %
    "toi":                 24,  # raw VND, flow
    "provision":           41,  # raw VND, flow (negative)
    "pbo":                 51,  # raw VND, flow
}


def load_paylater_channel_metrics():
    """Per-channel Paylater Q1.26 metrics from P&L_2026_Actual Paylater sheet.
    Channels: Total, VDS, VDS-EPASS, BE, VNPAY, FPT, MWG."""
    wb = openpyxl.load_workbook(RAW/"P&L_2026_Actual.xlsx", data_only=True)
    ws = wb["Paylater"]
    Q1_FLOW_COLS = [6, 7, 8]  # Jan, Feb, Mar
    Q1_END_COL = 8            # Mar 26 for state metrics
    DEC_COL = 4               # Dec 25 for new-issued calc
    out = {}
    for channel, base in PAYLATER_CHANNEL_BASES.items():
        d = {}
        for metric, off in PAYLATER_OFFSETS.items():
            row = base + off
            if metric in {"accumulated_account", "active_account", "balance", "lg2", "npl_cic", "npl_cake"}:
                v = ws.cell(row, Q1_END_COL).value
                d[metric] = v if isinstance(v, (int, float)) else None
            else:
                # Flow metric
                total = 0
                for c in Q1_FLOW_COLS:
                    v = ws.cell(row, c).value
                    if isinstance(v, (int, float)):
                        total += v
                d[metric] = total if total != 0 else None
        # Provision: store as abs (expense magnitude)
        if d.get("provision") is not None:
            d["provision"] = abs(d["provision"])
        # Compute new issued (accumulated Mar 26 - Dec 25)
        accum_dec = ws.cell(base + PAYLATER_OFFSETS["accumulated_account"], DEC_COL).value
        accum_mar = d.get("accumulated_account")
        if isinstance(accum_dec, (int, float)) and accum_mar is not None:
            d["new_issued_q1"] = accum_mar - accum_dec
        else:
            d["new_issued_q1"] = None
        # Compute % active rate
        if d.get("accumulated_account") and d.get("active_account"):
            d["active_rate"] = d["active_account"] / d["accumulated_account"]
        # Avg balance per active (proxy avg limit)
        if d.get("balance") and d.get("active_account"):
            d["avg_balance_per_active"] = d["balance"] / d["active_account"]
        # Spending per active (Q1 average) — using Mar value column 8
        spd_active_mar = ws.cell(base + 6, Q1_END_COL).value
        d["spending_per_active_mar"] = spd_active_mar if isinstance(spd_active_mar, (int, float)) else None
        out[channel] = d
    wb.close()
    return out


def load_paylater_q126_metrics():
    """Return Paylater Q1.26 detailed metrics from P&L_2026_Actual Paylater sheet.
    Cols: 3=Nov25, 4=Dec25, 5=YTD 2026, 6=Jan, 7=Feb, 8=Mar, 9=Apr."""
    wb = openpyxl.load_workbook(RAW/"P&L_2026_Actual.xlsx", data_only=True)
    ws = wb["Paylater"]
    def gv(r, c):
        v = ws.cell(r, c).value
        return v if isinstance(v, (int, float)) else None
    # End-of-Q1 = Mar 26 = col 8. Q1 flow = sum Jan+Feb+Mar = cols 6+7+8.
    accum_dec25 = gv(3, 4)  # Dec 25 accumulated account
    accum_mar26 = gv(3, 8)  # Mar 26 accumulated account
    active_mar26 = gv(4, 8)
    spending_q1 = sum(filter(None, [gv(8, 6), gv(8, 7), gv(8, 8)]))  # raw VND
    balance_mar26 = gv(9, 8)  # Principle balance Mar 26
    spending_per_accum_mar = gv(6, 8)
    spending_per_active_mar = gv(7, 8)
    new_issued_q1 = (accum_mar26 - accum_dec25) if (accum_mar26 and accum_dec25) else None
    active_rate_mar = (active_mar26 / accum_mar26) if (accum_mar26 and active_mar26) else None
    # Avg limit not directly in P&L file → use Balance / Active as proxy (approximation)
    avg_balance_per_active = (balance_mar26 / active_mar26) if (balance_mar26 and active_mar26) else None
    wb.close()
    return {
        "accumulated_account_eop": accum_mar26,           # accounts
        "active_account_eop": active_mar26,                # accounts
        "active_rate_eop": active_rate_mar,                # %
        "new_issued_q1": new_issued_q1,                    # accounts
        "spending_q1_bil_vnd": spending_q1 / 1e9 if spending_q1 else None,
        "balance_eop_bil_vnd": balance_mar26 / 1e9 if balance_mar26 else None,
        "spending_per_accum_account_vnd_mar": spending_per_accum_mar,    # VND per accum acct
        "spending_per_active_account_vnd_mar": spending_per_active_mar,  # VND per active acct
        "avg_balance_per_active_vnd_proxy": avg_balance_per_active,      # proxy for avg limit
    }


def derive_ue_q1_26(overall):
    """Q1.26 UE = sum Jan-Mar from existing P&L Actual data + Balance/NPL/LG2 from Total sheet."""
    def sum_q1(metric_arr):
        return sum(v for v in metric_arr[:3] if v is not None)
    # Pull Balance + NPL + LG2 directly from P&L Actual Total sheet
    wb = openpyxl.load_workbook(RAW/"P&L_2026_Actual.xlsx", data_only=True)
    ws = wb["Total"]
    # R9 Balance ENR (raw VND), monthly cols 6-9 = Jan-Apr 2026
    bal_dec25 = ws.cell(9, 4).value / 1e9 if ws.cell(9, 4).value else None  # Dec 25 EOP col 4
    bal_mar26 = ws.cell(9, 8).value / 1e9 if ws.cell(9, 8).value else None  # Mar 26 EOP col 8
    anr_q1 = (bal_dec25 + bal_mar26) / 2 if (bal_dec25 and bal_mar26) else None
    # R15 %NPL CAKE, R16 %NPL CIC, R14 % Attention debts #2 — all Mar 26 (col 8)
    npl_cake_mar26 = ws.cell(15, 8).value
    npl_cic_mar26 = ws.cell(16, 8).value
    lg2_mar26 = ws.cell(14, 8).value
    wb.close()
    return {
        "disb":      sum_q1(overall["2026 Actual"]["disb"]),
        "toi":       sum_q1(overall["2026 Actual"]["toi"]),
        "provision": sum_q1(overall["2026 Actual"]["prov"]),
        "pbo":       sum_q1(overall["2026 Actual"]["pbo"]),
        "pbt":       sum_q1(overall["2026 Actual"]["pbt"]),
        "balance":   anr_q1,  # ANR = avg(Dec25, Mar26) EOP balance
        "opex":      sum_q1(overall["2026 Actual"]["toi"]) - sum_q1(overall["2026 Actual"]["prov"]) - sum_q1(overall["2026 Actual"]["pbo"]),
        "npl_cake":  npl_cake_mar26 if isinstance(npl_cake_mar26, (int, float)) else None,
        "npl_cic":   npl_cic_mar26 if isinstance(npl_cic_mar26, (int, float)) else None,
        "lg2":       lg2_mar26 if isinstance(lg2_mar26, (int, float)) else None,
    }


def render_unit_economics(overall):
    q34 = load_ue_q3q4_25()
    q1 = derive_ue_q1_26(overall)
    quarters = ["Q3.25", "Q4.25", "Q1.26"]
    data = {"Q3.25": q34["Q3.25"], "Q4.25": q34["Q4.25"], "Q1.26": q1}

    parts = []
    parts.append("<p class='note'>Unit Economics tracking — 3 layers: <strong>5.1 Aggregate Lending</strong> (UPL quarterly Tỷ VND), <strong>5.2 CL by Channel trend</strong> (per-loan UE mVND, Q3.25 vs Q4.25), <strong>5.3 Q1.26 CL channel inputs</strong> (dropdown selector). Sources: <code>Unit Economics Q42025 Lending.xlsx</code> SUM/CL sheets + <code>CashLoan_Q1_26.xlsx</code> DATA sheet + P&L Actual file.</p>")

    # ===== Section 5.1 — Absolute values =====
    parts.append("<h3>5.1 — Aggregate Lending values per quarter (Tỷ VND)</h3>")
    parts.append("<table class='data'><thead><tr><th>Metric</th>" + "".join(f"<th>{q}</th>" for q in quarters) + "<th>Δ Q4 vs Q3</th><th>Δ Q1.26 vs Q4.25</th></tr></thead><tbody>")
    metric_rows = [
        ("disb", "Disbursement"),
        ("balance", "Balance (ANR)"),
        ("toi", "TOI"),
        ("provision", "Provision (expense)"),
        ("opex", "OPEX"),
        ("pbt", "PBT"),
    ]
    for key, label in metric_rows:
        vals = [data[q].get(key) for q in quarters]
        d_q4 = (vals[1] - vals[0]) if (vals[0] is not None and vals[1] is not None) else None
        d_q4_pct = (d_q4 / vals[0] * 100) if (d_q4 is not None and vals[0]) else None
        d_q1 = (vals[2] - vals[1]) if (vals[1] is not None and vals[2] is not None) else None
        d_q1_pct = (d_q1 / vals[1] * 100) if (d_q1 is not None and vals[1]) else None
        is_expense = key in {"provision", "opex"}
        parts.append("<tr><td class='label'>" + label + "</td>")
        for v in vals:
            parts.append(f"<td>{fmt(v)}</td>")
        d_q4_s = f"{fmt(d_q4)} ({d_q4_pct:+.0f}%)" if d_q4 is not None else "-"
        d_q1_s = f"{fmt(d_q1)} ({d_q1_pct:+.0f}%)" if d_q1 is not None else "-"
        cls_q4 = cell_class(d_q4, neg_bad=is_expense) if not is_expense else (("neg" if d_q4 and d_q4 > 0 else "") if d_q4 else "")
        cls_q1 = cell_class(d_q1, neg_bad=is_expense) if not is_expense else (("neg" if d_q1 and d_q1 > 0 else "") if d_q1 else "")
        parts.append(f"<td class='{cls_q4}'>{d_q4_s}</td><td class='{cls_q1}'>{d_q1_s}</td></tr>")
    # NPL CAKE + NPL CIC + LG2
    for key, label in [("npl_cake", "NPL CAKE %"), ("npl_cic", "NPL CIC %"), ("lg2", "Loan group 2 %")]:
        vals = [data[q].get(key) for q in quarters]
        parts.append("<tr><td class='label'>" + label + "</td>")
        for v in vals:
            parts.append(f"<td>{(f'{v*100:.2f}%' if v is not None else '-')}</td>")
        parts.append("<td>-</td><td>-</td></tr>")
    parts.append("</tbody></table>")

    # ===== Section 5.1 ratios bundled (formerly 5.2/5.3) =====
    parts.append("<h4>Ratios per ANR (quarterly rate)</h4>")
    parts.append("<table class='data'><thead><tr><th>Ratio</th>" + "".join(f"<th>{q}</th>" for q in quarters) + "</tr></thead><tbody>")
    anr_ratios = [
        ("toi", "TOI / ANR  (yield)"),
        ("provision", "Provision / ANR  (CoR)"),
        ("opex", "OPEX / ANR"),
        ("pbt", "PBT / ANR  (return)"),
    ]
    for key, label in anr_ratios:
        parts.append("<tr><td class='label'>" + label + "</td>")
        for q in quarters:
            anr = data[q].get("balance")
            v = data[q].get(key)
            r = (v / anr * 100) if (anr and v is not None) else None
            parts.append(f"<td>{(f'{r:.1f}%' if r is not None else '-')}</td>")
        parts.append("</tr>")
    parts.append("</tbody></table>")

    # ===== Section 5.1 ratios /Disb =====
    parts.append("<h4>Ratios per Disbursement</h4>")
    parts.append("<table class='data'><thead><tr><th>Ratio</th>" + "".join(f"<th>{q}</th>" for q in quarters) + "</tr></thead><tbody>")
    disb_ratios = [
        ("toi", "TOI / Disb"),
        ("provision", "Provision / Disb"),
        ("opex", "OPEX / Disb"),
        ("pbt", "PBT / Disb"),
    ]
    for key, label in disb_ratios:
        parts.append("<tr><td class='label'>" + label + "</td>")
        for q in quarters:
            disb = data[q].get("disb")
            v = data[q].get(key)
            r = (v / disb * 100) if (disb and v is not None) else None
            parts.append(f"<td>{(f'{r:.1f}%' if r is not None else '-')}</td>")
        parts.append("</tr>")
    parts.append("</tbody></table>")

    # ===== Section 5.1.1 — TOI / Risk / Net Yield focus view (channel-filterable) =====
    import json as _json
    cohort_focus = load_cohort_ue_per_channel()
    parts.append("<h3>5.1.1 — TOI / Risk / Net Yield by Product — %ANR &amp; %Disbursement (channel filter)</h3>")
    parts.append("<p class='note'>Per-loan cohort UE. <strong>Risk = Provision</strong>. <strong>Net Yield = TOI − Risk</strong>. Chọn channel ở dropdown — mặc định <code>All</code> (total). Cost breakdown (OPEX components) skip ở view này.</p>")
    # Channel list across all products
    _all_ch = set()
    for _p in cohort_focus:
        for _q in cohort_focus[_p]:
            _all_ch.update(cohort_focus[_p][_q].keys())
    _ch_order = ["All", "CAKE", "VDS", "MWG", "VNPAY", "ZALOPAY", "VNPOST", "BE"]
    _channels_sorted = [c for c in _ch_order if c in _all_ch] + sorted(_all_ch - set(_ch_order))
    parts.append("<div style='margin:8px 0;'><label><strong>Channel:</strong> <select id='ue-focus-channel' onchange='renderUEFocus()' style='padding:4px 10px;'>")
    for ch in _channels_sorted:
        parts.append(f"<option value='{ch}'>{ch}</option>")
    parts.append("</select></label></div>")
    parts.append("<div id='ue-focus-tables'></div>")
    parts.append(f"<script>\nconst UE_FOCUS_DATA = {_json.dumps(cohort_focus)};\n")
    parts.append("""
const UE_FOCUS_QUARTERS = ["Q3.25","Q4.25","Q1.26"];
const UE_FOCUS_PRODUCTS = [["CL","CashLoan"],["PD","Payday"],["PL","Paylater"]];
function fmtNum(v,d){ return (v===null||v===undefined)?"-":v.toLocaleString("en-US",{minimumFractionDigits:d,maximumFractionDigits:d}); }
function fmtPct(v){ return (v===null||v===undefined)?"-":v.toFixed(1)+"%"; }
function renderUEFocus(){
  const ch = document.getElementById('ue-focus-channel').value;
  let html = "";
  for (const [pk,pl] of UE_FOCUS_PRODUCTS){
    html += "<h4>"+pl+" — channel: "+ch+"</h4>";
    html += "<table class='data'><thead><tr><th>Metric</th>";
    for (const q of UE_FOCUS_QUARTERS) html += "<th>"+q+" mVND</th><th>"+q+" %ANR</th><th>"+q+" %Disb</th>";
    html += "</tr></thead><tbody>";
    const rows = [
      ["TOI","toi_mvnd","toi_per_anr_pct","toi_per_disb_pct"],
      ["Risk (Provision)","provision_mvnd","provision_per_anr_pct","provision_per_disb_pct"],
      ["Net Yield (TOI−Risk)","net_yield_mvnd","net_yield_per_anr_pct","net_yield_per_disb_pct"]
    ];
    for (const [label,vk,ak,dk] of rows){
      html += "<tr><td class='label'>"+label+"</td>";
      for (const q of UE_FOCUS_QUARTERS){
        const e = (UE_FOCUS_DATA[pk]&&UE_FOCUS_DATA[pk][q]&&UE_FOCUS_DATA[pk][q][ch])?UE_FOCUS_DATA[pk][q][ch]:null;
        if(e){
          const cls = label.startsWith("Net Yield")?"accent":"";
          html += "<td class='"+cls+"'>"+fmtNum(e[vk],3)+"</td><td class='"+cls+"'>"+fmtPct(e[ak])+"</td><td class='"+cls+"'>"+fmtPct(e[dk])+"</td>";
        } else { html += "<td>-</td><td>-</td><td>-</td>"; }
      }
      html += "</tr>";
    }
    html += "</tbody></table>";
  }
  document.getElementById('ue-focus-tables').innerHTML = html;
}
renderUEFocus();
</script>""")

    # ===== Section 5.2 — By Product trend (aggregate Tỷ VND, Q3.25 → Q4.25 → Q1.26) =====
    parts.append("<h3>5.2 — By Product Trend — aggregate quarterly (Tỷ VND)</h3>")
    parts.append("<p class='legend'>Source: P&L_2025_final.xlsx (channels summed → product aggregate) + P&L_2026_Actual.xlsx product sheets. Products: CL, PD, OD. PL/CC pending.</p>")
    by_prod = load_by_product_quarterly()
    products = ["CL", "PD", "OD", "PL"]
    qts = ["Q3.25", "Q4.25", "Q1.26"]
    metrics_5_2 = [("disb","Disbursement (Tỷ)"), ("balance","Balance/ANR (Tỷ)"), ("toi","TOI (Tỷ)"), ("provision","Provision (Tỷ, abs)"), ("pbo","Profit before overhead (Tỷ)")]
    parts.append("<table class='data'><thead><tr><th>Product</th><th>Metric</th>" + "".join(f"<th>{q}</th>" for q in qts) + "<th>Δ Q4 vs Q3</th><th>Δ Q1 vs Q4</th><th>%Disb Q3</th><th>%Disb Q4</th><th>%Disb Q1</th></tr></thead><tbody>")
    for product in products:
        for i, (key, label) in enumerate(metrics_5_2):
            vals = [by_prod[product][q].get(key) for q in qts]
            disbs = [by_prod[product][q].get("disb") for q in qts]
            d_q4 = (vals[1] - vals[0]) if (vals[0] is not None and vals[1] is not None) else None
            d_q4_pct = (d_q4 / vals[0] * 100) if (d_q4 is not None and vals[0]) else None
            d_q1 = (vals[2] - vals[1]) if (vals[1] is not None and vals[2] is not None) else None
            d_q1_pct = (d_q1 / vals[1] * 100) if (d_q1 is not None and vals[1]) else None
            ratios = []
            for v, d in zip(vals, disbs):
                if v is not None and d:
                    ratios.append(f"{v/d*100:.1f}%")
                else:
                    ratios.append("-")
            row_html = f"<tr><td class='label'>{product if i==0 else ''}</td><td class='label'>{label}</td>"
            for v in vals: row_html += f"<td>{fmt(v)}</td>"
            d_q4_s = f"{fmt(d_q4)} ({d_q4_pct:+.0f}%)" if d_q4 is not None else "-"
            d_q1_s = f"{fmt(d_q1)} ({d_q1_pct:+.0f}%)" if d_q1 is not None else "-"
            row_html += f"<td>{d_q4_s}</td><td>{d_q1_s}</td>"
            for r in ratios: row_html += f"<td>{r}</td>"
            row_html += "</tr>"
            parts.append(row_html)
    parts.append("</tbody></table>")

    # NPL + LG2 per product across quarters
    parts.append("<h4>Risk metrics — NPL CAKE + NPL CIC + Loan group 2 by Product (state at end of quarter)</h4>")
    parts.append("<table class='data'><thead><tr><th>Product</th><th>Metric</th>" + "".join(f"<th>{q}</th>" for q in qts) + "<th>Trend</th></tr></thead><tbody>")
    for product in products:
        risk_metrics = [("npl_cake", "NPL CAKE %"), ("npl_cic", "NPL CIC %"), ("lg2", "Loan group 2 %")]
        for idx, (risk_key, risk_label) in enumerate(risk_metrics):
            risk_vals = [by_prod[product][q].get(risk_key) for q in qts]
            row_html = f"<tr><td class='label'>{product if idx==0 else ''}</td><td class='label'>{risk_label}</td>"
            for v in risk_vals:
                row_html += f"<td>{(f'{v*100:.2f}%' if v is not None else '-')}</td>"
            valid = [v for v in risk_vals if v is not None]
            if len(valid) >= 2:
                trend = "↗ tăng" if valid[-1] > valid[0] else "↘ giảm" if valid[-1] < valid[0] else "→ flat"
                row_html += f"<td>{trend}</td>"
            else:
                row_html += "<td>-</td>"
            row_html += "</tr>"
            parts.append(row_html)
    parts.append("</tbody></table>")

    # ===== Section 5.2.1 — CL by Channel cohort UE (per-loan, Q3.25 vs Q4.25) =====
    parts.append("<h3>5.2.1 — CL Cohort UE by Channel — per-loan view (mVND/loan, %ANR)</h3>")
    parts.append("<p class='legend'>Source: Q425 SUM file CL sheet — per-loan lifetime cohort UE (model output). <strong>ANR ở đây là per-loan balance (mVND/loan), không phải aggregate Tỷ VND.</strong> Q1.26 cohort UE per channel chưa available (Leadsheet trong Q126 file chỉ chứa Cake cohort). Channel P&L Q1.26 xem Section 4 (deep dive).</p>")
    cl_ue = load_cl_per_loan_ue()
    cl_metrics = [
        ("anr", "ANR", "mVND/loan"),
        ("toi", "TOI", "mVND/loan"),
        ("nii", "Net Interest Income", "mVND/loan"),
        ("nfi", "Net Fee Income", "mVND/loan"),
        ("provision", "Provision", "mVND/loan"),
        ("opex", "OPEX (total)", "mVND/loan"),
        ("acquisition", "  Acquisition cost", "mVND/loan"),
        ("uw_disbursal", "  UW & Disbursal", "mVND/loan"),
        ("collection", "  Collection", "mVND/loan"),
        ("aftersale", "  Aftersale service", "mVND/loan"),
        ("support_unit", "  Support unit", "mVND/loan"),
        ("pbt", "PBT", "mVND/loan"),
        ("pbsc", "PBSC (PBT before Support)", "mVND/loan"),
    ]
    cl_channels = ["CL all", "CAKE", "VDS", "MWG", "VNPAY", "ZLP"]
    for view in ["value", "pct"]:
        view_label = "Absolute (mVND/loan)" if view == "value" else "% ANR (lifetime ratio)"
        parts.append(f"<h4>{view_label}</h4>")
        parts.append("<table class='data'><thead><tr><th>Metric</th>" +
                     "".join(f"<th>{ch}<br><small>Q3.25</small></th><th>{ch}<br><small>Q4.25</small></th>" for ch in cl_channels) + "</tr></thead><tbody>")
        for key, label, _ in cl_metrics:
            row = f"<tr><td class='label'>{label}</td>"
            for ch in cl_channels:
                v_q3, p_q3 = cl_ue[ch]["Q3.25"][key]
                v_q4, p_q4 = cl_ue[ch]["Q4.25"][key]
                if view == "value":
                    row += f"<td>{(f'{v_q3:.3f}' if v_q3 is not None else '-')}</td>"
                    row += f"<td>{(f'{v_q4:.3f}' if v_q4 is not None else '-')}</td>"
                else:
                    row += f"<td>{(f'{p_q3*100:.1f}%' if p_q3 is not None else '-')}</td>"
                    row += f"<td>{(f'{p_q4*100:.1f}%' if p_q4 is not None else '-')}</td>"
            row += "</tr>"
            parts.append(row)
        parts.append("</tbody></table>")

    # ===== Section 5.2.5 — Cohort UE per Channel (all products, Q3.25 + Q4.25) =====
    parts.append("<h3>5.2.5 — Cohort UE per Channel — lifetime per-loan (mVND/loan)</h3>")
    parts.append("<p class='note'>Per-loan lifetime cohort UE từ <strong>file riêng từng channel</strong> trong Q325 + Q425 folders. Coverage: CL/PD/PL × các channel có file. <strong>Q1.26 chỉ có CAKE</strong> — file Q126 là combined file, dropdown sub-segment khóa ở Cake; muốn channel khác phải mở file đổi dropdown + save riêng.</p>")
    cohort_ue = load_cohort_ue_per_channel()
    for product in ["CL", "PD", "PL"]:
        if product not in cohort_ue:
            continue
        parts.append(f"<h4>{product} — cohort UE by channel</h4>")
        parts.append("<table class='data'><thead><tr><th>Channel</th><th>Quarter</th>"
                     "<th>ANR (mVND)</th><th>TOI (mVND)</th><th>Provision (mVND)</th><th>OPEX (mVND)</th><th>PBT (mVND)</th>"
                     "<th>Net Yield (mVND)</th><th>TOI/ANR</th><th>Prov/ANR</th><th>PBT/ANR</th></tr></thead><tbody>")
        # Gather all channels across quarters
        all_channels = set()
        for q in cohort_ue[product]:
            all_channels.update(cohort_ue[product][q].keys())
        for channel in sorted(all_channels):
            for quarter in ["Q3.25", "Q4.25", "Q1.26"]:
                m = cohort_ue[product].get(quarter, {}).get(channel)
                if not m:
                    continue
                def mv(v, suf=""):
                    return (f"{v:.3f}{suf}" if v is not None else "-")
                def pct(v):
                    return (f"{v:.1f}%" if v is not None else "-")
                parts.append(f"<tr><td class='label'>{channel}</td><td class='label'>{quarter}</td>"
                             f"<td>{mv(m.get('anr_mvnd'))}</td><td>{mv(m.get('toi_mvnd'))}</td>"
                             f"<td class='neg'>{mv(m.get('provision_mvnd'))}</td><td class='neg'>{mv(m.get('opex_mvnd'))}</td>"
                             f"<td class='accent'>{mv(m.get('pbt_mvnd'))}</td>"
                             f"<td class='accent'>{mv(m.get('net_yield_mvnd'))}</td>"
                             f"<td>{pct(m.get('toi_per_anr_pct'))}</td><td>{pct(m.get('provision_per_anr_pct'))}</td>"
                             f"<td>{pct(m.get('pbt_per_anr_pct'))}</td></tr>")
        parts.append("</tbody></table>")

    # ===== Section 5.2.2 — Term Loan pricing inputs Q1.26 (CL + PD aggregate) =====
    parts.append("<h3>5.2.2 — Term Loan pricing inputs Q1.26 (CL + PD aggregate, all-channels)</h3>")
    parts.append("<p class='legend'>Aggregate-level pricing/funnel inputs cho 2 term loan products Q1.26. Source: UE Q126 DATA sheet 'All' row. Per-channel breakdown trong 5.3.</p>")
    cl_inputs_all = load_q126_cl_channel_inputs().get("All", {})
    pd_inputs_all = load_q126_pd_channel_inputs().get("All", {})
    parts.append("<table class='data'><thead><tr><th>Metric</th><th>CL aggregate</th><th>PD aggregate</th></tr></thead><tbody>")
    pricing_rows = [
        ("avg_ticket", "AVG Ticket size (mVND)", lambda v: f"{v:.2f}" if v else "-"),
        ("avg_tenor", "AVG Tenor (months)", lambda v: f"{v:.0f}" if v else "-"),
        ("ir", "IR (interest rate)", lambda v: f"{v*100:.2f}%" if v else "-"),
        ("cof", "COF (cost of funds)", lambda v: f"{v*100:.2f}%" if v else "-"),
        ("nim_calc", "NIM (= IR − COF)", lambda v: f"{v*100:.2f}%" if v else "-"),
        ("approval_rate", "Approval rate", lambda v: f"{v*100:.1f}%" if v else "-"),
        ("signed_rate", "Signed rate", lambda v: f"{v*100:.1f}%" if v else "-"),
        ("insurance_pen", "Insurance penetration", lambda v: f"{v*100:.1f}%" if v else "-"),
    ]
    cl_inputs_all["nim_calc"] = (cl_inputs_all.get("ir") or 0) - (cl_inputs_all.get("cof") or 0) if cl_inputs_all.get("ir") else None
    pd_inputs_all["nim_calc"] = (pd_inputs_all.get("ir") or 0) - (pd_inputs_all.get("cof") or 0) if pd_inputs_all.get("ir") else None
    for key, label, fmt_fn in pricing_rows:
        parts.append(f"<tr><td class='label'>{label}</td>"
                     f"<td>{fmt_fn(cl_inputs_all.get(key))}</td>"
                     f"<td>{fmt_fn(pd_inputs_all.get(key))}</td></tr>")
    parts.append("</tbody></table>")

    # ===== Section 5.2.3 — Paylater specific metrics Q1.26 =====
    parts.append("<h3>5.2.3 — Paylater specific metrics Q1.26 (Top-level, all-channel)</h3>")
    parts.append("<p class='legend'>Source: P&L_2026_Actual.xlsx Paylater sheet. State metrics = end-of-Q1 (Mar 26). Flow metrics = sum Jan+Feb+Mar. Avg limit chính xác chưa có trong P&L file — dùng Balance/Active làm proxy.</p>")
    pl_m = load_paylater_q126_metrics()
    parts.append("<table class='data'><thead><tr><th>Metric</th><th>Q1.26</th><th>Notes</th></tr></thead><tbody>")
    paylater_rows = [
        ("New issued (Q1 = Δ accumulated)", pl_m.get("new_issued_q1"), "{:,.0f} accounts", "= Accum Mar 26 − Accum Dec 25"),
        ("Lũy kế issued / Accumulated account (EOP Mar 26)", pl_m.get("accumulated_account_eop"), "{:,.0f} accounts", "Total Paylater accounts ever issued"),
        ("Active account (EOP Mar 26)", pl_m.get("active_account_eop"), "{:,.0f} accounts", "Active in current quarter"),
        ("% Active rate (EOP Mar 26)", pl_m.get("active_rate_eop"), "{:.1%}", "Active / Accumulated"),
        ("Spending Q1 (Jan+Feb+Mar)", pl_m.get("spending_q1_bil_vnd"), "{:,.1f} Tỷ VND", "Total spend trong Q1"),
        ("Balance EOP Mar 26 (Principle)", pl_m.get("balance_eop_bil_vnd"), "{:,.1f} Tỷ VND", "Outstanding principle balance"),
        ("Spending / Active account (Mar)", pl_m.get("spending_per_active_account_vnd_mar"), "{:,.0f} VND", "Spend per active account in March"),
        ("Spending / Accumulated account (Mar)", pl_m.get("spending_per_accum_account_vnd_mar"), "{:,.0f} VND", "Spend per accum account in March"),
        ("Avg Balance / Active (proxy for avg limit)", pl_m.get("avg_balance_per_active_vnd_proxy"), "{:,.0f} VND", "Proxy — file P&L không có direct Limit metric"),
    ]
    for label, val, fmt_str, note in paylater_rows:
        if val is None:
            val_s = "-"
        else:
            try:
                val_s = fmt_str.format(val)
            except (ValueError, TypeError):
                val_s = str(val)
        parts.append(f"<tr><td class='label'>{label}</td><td>{val_s}</td><td><small>{note}</small></td></tr>")
    parts.append("</tbody></table>")

    # ===== Section 5.2.4 — Paylater per-channel Q1.26 =====
    parts.append("<h3>5.2.4 — Paylater Channel breakdown Q1.26</h3>")
    parts.append("<p class='legend'>Source: P&L_2026_Actual Paylater sheet, channel sub-sections (VDS R54, VDS-EPASS R107, BE R160, VNPAY R213, FPT R266, MWG R319). State = Mar 26 EOP. Flow = sum Jan+Feb+Mar.</p>")
    pl_ch = load_paylater_channel_metrics()
    pl_channels = ["Total", "VDS", "VDS-EPASS", "BE", "VNPAY", "FPT", "MWG"]
    parts.append("<table class='data'><thead><tr><th>Metric</th>" + "".join(f"<th>{ch}</th>" for ch in pl_channels) + "</tr></thead><tbody>")
    pl_rows = [
        ("New issued Q1 (Δ accum)", "new_issued_q1", "{:,.0f}", 1),
        ("Lũy kế Accumulated (Mar 26)", "accumulated_account", "{:,.0f}", 1),
        ("Active account (Mar 26)", "active_account", "{:,.0f}", 1),
        ("% Active rate", "active_rate", "{:.1%}", 1),
        ("Spending Q1 (Tỷ VND)", "spending", "{:,.1f}", 1e9),
        ("Balance EOP Mar 26 (Tỷ VND)", "balance", "{:,.1f}", 1e9),
        ("TOI Q1 (Tỷ VND)", "toi", "{:,.2f}", 1e9),
        ("Provision Q1 (Tỷ VND, abs)", "provision", "{:,.2f}", 1e9),
        ("PBO Q1 (Tỷ VND)", "pbo", "{:,.2f}", 1e9),
        ("Spending/Active per month (VND, Mar)", "spending_per_active_mar", "{:,.0f}", 1),
        ("Avg Balance/Active (VND, proxy avg limit)", "avg_balance_per_active", "{:,.0f}", 1),
        ("NPL CAKE % (Mar 26)", "npl_cake", "{:.2%}", 1),
        ("NPL CIC % (Mar 26)", "npl_cic", "{:.2%}", 1),
    ]
    for label, key, fmt_str, divisor in pl_rows:
        row_html = f"<tr><td class='label'>{label}</td>"
        for ch in pl_channels:
            v = pl_ch.get(ch, {}).get(key)
            if v is None:
                row_html += "<td>-</td>"
            else:
                try:
                    val = v / divisor if divisor != 1 else v
                    row_html += f"<td>{fmt_str.format(val)}</td>"
                except (ValueError, TypeError):
                    row_html += f"<td>{v}</td>"
        row_html += "</tr>"
        parts.append(row_html)
    parts.append("</tbody></table>")
    parts.append("<p class='legend'><strong>Note:</strong> LG2 % và một số NPL cells trong source file = 0 (bug data) → mình hiển thị NPL CAKE và CIC từ file gốc, LG2 chưa available per Paylater channel.</p>")

    # ===== Section 5.3 — Q1.26 Channel Snapshot with dropdown (CL + PD) =====
    parts.append("<h3>5.3 — Q1.26 Channel snapshot (CL + PD, with dropdown)</h3>")
    parts.append("<p class='legend'>Pricing/funnel inputs từ Q126 DATA sheets per channel. CL = 9 channels, PD = 5 channels. Dropdown highlight 1 channel.</p>")
    q1_inputs = load_q126_cl_channel_inputs()
    q1_channels = list(Q126_CL_DATA_ROWS.keys())
    parts.append("<div style='margin: 8px 0;'><label><strong>Highlight channel:</strong> <select id='ue-channel-select' onchange='highlightChannel()' style='padding:4px 8px;'>")
    parts.append("<option value=''>(all channels)</option>")
    for ch in q1_channels:
        parts.append(f"<option value='{ch}'>{ch}</option>")
    parts.append("</select></label></div>")
    parts.append("<h4>CashLoan channels Q1.26</h4>")
    parts.append("<table class='data ue-q126-cl'><thead><tr><th>Channel</th><th>IR</th><th>COF</th><th>NIM (IR-COF)</th><th>AVG Ticket (mVND)</th><th>AVG Tenor (mo)</th><th>Approval rate</th><th>Signed rate</th><th>Insurance penetration</th></tr></thead><tbody>")
    for ch in q1_channels:
        d = q1_inputs[ch]
        ir = d.get("ir") or 0
        cof = d.get("cof") or 0
        nim = ir - cof if (ir and cof) else None
        ticket = d.get("avg_ticket")
        tenor = d.get("avg_tenor")
        appr = d.get("approval_rate")
        signed = d.get("signed_rate")
        ins_pen = d.get("insurance_pen")
        ir_s = f"{ir*100:.2f}%" if ir else "-"
        cof_s = f"{cof*100:.2f}%" if cof else "-"
        nim_s = f"{nim*100:.2f}%" if nim is not None else "-"
        ticket_s = f"{ticket:.2f}" if ticket else "-"
        tenor_s = f"{tenor:.0f}" if tenor else "-"
        appr_s = f"{appr*100:.1f}%" if appr else "-"
        signed_s = f"{signed*100:.1f}%" if signed else "-"
        ins_pen_s = f"{ins_pen*100:.1f}%" if ins_pen else "-"
        parts.append(f"<tr data-channel='{ch}'><td class='label'>{ch}</td>"
                     f"<td>{ir_s}</td><td>{cof_s}</td><td>{nim_s}</td>"
                     f"<td>{ticket_s}</td><td>{tenor_s}</td>"
                     f"<td>{appr_s}</td><td>{signed_s}</td><td>{ins_pen_s}</td></tr>")
    parts.append("</tbody></table>")

    # PD channels Q1.26
    parts.append("<h4>Payday channels Q1.26</h4>")
    pd_inputs = load_q126_pd_channel_inputs()
    pd_channels = list(Q126_PD_DATA_ROWS.keys())
    parts.append("<table class='data ue-q126-cl'><thead><tr><th>Channel</th><th>IR</th><th>COF</th><th>NIM (IR-COF)</th><th>AVG Ticket (mVND)</th><th>AVG Tenor (mo)</th><th>Approval rate</th><th>Signed rate</th><th>Insurance penetration</th></tr></thead><tbody>")
    for ch in pd_channels:
        d = pd_inputs[ch]
        ir = d.get("ir") or 0
        cof = d.get("cof") or 0
        nim = ir - cof if (ir and cof) else None
        ticket = d.get("avg_ticket")
        tenor = d.get("avg_tenor")
        appr = d.get("approval_rate")
        signed = d.get("signed_rate")
        ins_pen = d.get("insurance_pen")
        parts.append(f"<tr data-channel='{ch}'><td class='label'>{ch}</td>"
                     f"<td>{(f'{ir*100:.2f}%' if ir else '-')}</td>"
                     f"<td>{(f'{cof*100:.2f}%' if cof else '-')}</td>"
                     f"<td>{(f'{nim*100:.2f}%' if nim is not None else '-')}</td>"
                     f"<td>{(f'{ticket:.2f}' if ticket else '-')}</td>"
                     f"<td>{(f'{tenor:.0f}' if tenor else '-')}</td>"
                     f"<td>{(f'{appr*100:.1f}%' if appr else '-')}</td>"
                     f"<td>{(f'{signed*100:.1f}%' if signed else '-')}</td>"
                     f"<td>{(f'{ins_pen*100:.1f}%' if ins_pen else '-')}</td></tr>")
    parts.append("</tbody></table>")
    parts.append("""<script>
function highlightChannel(){
  var sel = document.getElementById('ue-channel-select').value;
  var rows = document.querySelectorAll('table.ue-q126-cl tbody tr');
  rows.forEach(function(r){
    if (sel === '' || r.getAttribute('data-channel') === sel){ r.style.opacity='1'; r.style.background=''; if(sel===r.getAttribute('data-channel')) r.style.background='#fff3cd'; } else { r.style.opacity='0.3'; r.style.background=''; }
  });
}
</script>""")

    # Quick insights
    parts.append("<h3>5.4 — Trend reads</h3>")
    parts.append("<ol class='insights'>")
    # Compute ratios for insights
    def ratio(q, num, denom):
        v = data[q].get(num); d = data[q].get(denom)
        return (v/d*100) if (v is not None and d) else None
    yld = [ratio(q, "toi", "disb") for q in quarters]
    cor = [ratio(q, "provision", "disb") for q in quarters]
    pbt_disb = [ratio(q, "pbt", "disb") for q in quarters]
    parts.append(f"<li><strong>Yield / Disbursement</strong>: Q3.25={yld[0]:.1f}% → Q4.25={yld[1]:.1f}% → Q1.26={yld[2]:.1f}%. " +
                 ("Yield expanding mạnh — pricing power tăng / mix shift sang segment yield cao." if yld[2] > yld[0] else "Yield compressing — cần check IR pressure hoặc mix shift sang segment low-yield.") + "</li>")
    parts.append(f"<li><strong>Cost of Risk (Provision/Disb)</strong>: Q3.25={cor[0]:.1f}% → Q4.25={cor[1]:.1f}% → Q1.26={cor[2]:.1f}%. " +
                 ("CoR đang tăng — NPL pressure phát sinh, segment risk shift hoặc booking acceleration outpace recovery." if cor[2] > cor[0] else "CoR đang giảm — credit quality cải thiện hoặc collection hiệu quả hơn.") + "</li>")
    parts.append(f"<li><strong>PBT margin (PBT/Disb)</strong>: Q3.25={pbt_disb[0]:.1f}% → Q4.25={pbt_disb[1]:.1f}% → Q1.26={pbt_disb[2]:.1f}%. " +
                 (f"PBT margin Q1.26 đã <strong>vượt mạnh</strong> 2 quý trước." if pbt_disb[2] > max(pbt_disb[0], pbt_disb[1]) else "Margin nén lại cần điều tra.") + "</li>")
    nc = [data[q].get("npl_cake") for q in quarters]
    if nc[0] is not None and nc[1] is not None:
        npl_trend = "giảm" if nc[1] < nc[0] else "tăng"
        q1_str = f" → Q1.26={nc[2]*100:.2f}%" if nc[2] is not None else ""
        parts.append(f"<li><strong>NPL CAKE</strong>: Q3.25={nc[0]*100:.2f}% → Q4.25={nc[1]*100:.2f}% ({npl_trend}){q1_str}. NPL CIC xem section 5.2 (UE SUM file Q3/Q4 không tách CIC).</li>")
    parts.append("</ol>")

    return "\n".join(parts)


# ========== Main HTML wrapper ==========
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<title>Cake Lending — P&L & Unit Economics</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Be+Vietnam+Pro:wght@400;600;700&family=JetBrains+Mono:wght@400;600&display=swap" rel="stylesheet">
<style>
  /* === Cake Brand Theme === */
  :root {{
    --pink-500: #FF2D71; --pink-700: #B8154F;
    --grape-500: #7B2FBE; --grape-900: #1A1033;
    --mint-500: #0BC76A; --amber-500: #FF7A2F;
    --surface-1: #F7F7F9; --surface-2: #FFFFFF; --page-bg: #E8E8EF;
  }}
  body {{ font-family: 'Be Vietnam Pro', -apple-system, system-ui, sans-serif; max-width: 1600px;
         margin: 0 auto; padding: 16px; color: var(--grape-900); line-height: 1.45; background: var(--page-bg); }}
  h1 {{ font-weight: 700; border-bottom: 3px solid var(--pink-500); padding-bottom: 8px; margin-bottom: 4px; color: var(--grape-900); }}
  h2 {{ font-weight: 600; background: var(--surface-2); padding: 10px 16px; border-left: 4px solid var(--pink-500);
        border-radius: 0 12px 12px 0; margin-top: 28px; color: var(--grape-900); }}
  h3 {{ font-weight: 600; color: var(--grape-500); margin-top: 24px; }}
  h4 {{ font-weight: 600; color: var(--pink-700); margin-top: 16px; margin-bottom: 6px; }}
  .meta {{ color: #6b6478; font-size: 0.9em; margin-bottom: 16px; }}
  .note {{ background: #FFF1E8; border-left: 3px solid var(--amber-500); padding: 8px 12px; font-size: 0.9em; border-radius: 0 8px 8px 0; }}
  table.data {{ border-collapse: collapse; width: 100%; margin: 12px 0; font-size: 0.85em; background: var(--surface-2);
               border-radius: 12px; overflow: hidden; box-shadow: 0 1px 4px rgba(26,16,51,0.06); }}
  table.data th, table.data td {{ border: 1px solid #E4E1EC; padding: 6px 9px; text-align: right; }}
  table.data th {{ background: var(--grape-500); color: white; font-weight: 600; }}
  table.data td.label {{ text-align: left; font-weight: 600; background: var(--surface-1); }}
  table.data td.accent {{ background: #FFF1E8; font-weight: 600; }}
  table.data th.accent {{ background: var(--amber-500); }}
  table.data tr.delta td {{ background: var(--surface-1); font-style: italic; }}
  table.data tr.total-row td {{ background: #FFF1E8; font-weight: 700; border-top: 2px solid var(--pink-500); }}
  td.neg {{ color: var(--pink-700); font-weight: 600; }}
  td.pos-neg {{ color: var(--pink-700); font-weight: 600; }}
  td.pos {{ color: var(--mint-500); font-weight: 600; }}
  .na {{ color: #A8A2B8; }}
  .legend {{ font-size: 0.85em; color: #6b6478; margin: 8px 0; }}
  ol.insights {{ background: var(--surface-2); padding: 16px 16px 16px 36px; border-left: 4px solid var(--grape-500); border-radius: 0 12px 12px 0; }}
  ol.insights li {{ margin: 8px 0; line-height: 1.5; }}
  code, .mono {{ font-family: 'JetBrains Mono', monospace; background: #EEEAF6; padding: 1px 5px; border-radius: 4px; font-size: 0.88em; }}
  /* Tabs */
  .tab-nav {{ display: flex; gap: 6px; border-bottom: 3px solid var(--pink-500); margin: 16px 0 0; }}
  .tab-btn {{ font-family: 'Be Vietnam Pro', sans-serif; padding: 11px 26px; background: var(--surface-1);
              border: none; border-radius: 12px 12px 0 0; font-size: 1em; font-weight: 600; cursor: pointer; color: #6b6478; }}
  .tab-btn:hover {{ background: #DCD7E8; }}
  .tab-btn.active {{ background: linear-gradient(135deg, var(--pink-500), var(--grape-500)); color: white;
                     box-shadow: 0 4px 14px rgba(255,45,113,0.28); }}
  .tab-content {{ display: none; padding: 8px 0; }}
  .tab-content.active {{ display: block; }}
  select {{ font-family: 'Be Vietnam Pro', sans-serif; border: 1px solid var(--grape-500); border-radius: 8px; }}
</style>
</head>
<body>
<h1>Cake Lending — P&L &amp; Unit Economics</h1>
<p class="meta">Generated by <code>lending_pl_html_report.py</code> · Đơn vị: <strong>Tỷ VND</strong> · 2026 Actual YTD = Jan–Apr · UE quarterly Q3.25 → Q1.26</p>

<div class="tab-nav">
  <button class="tab-btn active" onclick="showTab(event,'tab-pnl-2025')">1 · P&L 2025</button>
  <button class="tab-btn" onclick="showTab(event,'tab-pnl-2026')">2 · P&L 2026</button>
  <button class="tab-btn" onclick="showTab(event,'tab-unit-econ')">3 · Unit Economics</button>
</div>

<div id="tab-pnl-2025" class="tab-content active">
{tab_pnl_2025}
</div>

<div id="tab-pnl-2026" class="tab-content">
{tab_pnl_2026}
</div>

<div id="tab-unit-econ" class="tab-content">
{tab_unit_econ}
</div>

<script>
function showTab(evt, tabId) {{
  document.querySelectorAll('.tab-content').forEach(function(c){{ c.classList.remove('active'); }});
  document.querySelectorAll('.tab-btn').forEach(function(b){{ b.classList.remove('active'); }});
  document.getElementById(tabId).classList.add('active');
  evt.currentTarget.classList.add('active');
}}
</script>
</body>
</html>
"""


Q425_CL_DIR = RAW.parent / "Unit Economic" / "Q425" / "Cash Loan"
Q425_PD_DIR = RAW.parent / "Unit Economic" / "Q425" / "Payday"
# Sharing params live in Leadsheet R47-R54, value col 5
SHARING_PARAM_ROWS = {
    "sharing_toi_risk_dpd90": 47,
    "sharing_toi_risk_dpd180": 48,
    "sharing_cii_only": 50,
    "sharing_cof_only": 51,
    "sharing_nfi_only": 52,
    "sharing_risk_dpd90_only": 53,
    "sharing_risk_dpd180_only": 54,
}
# filename → channel for Q425 channel files
Q425_CL_FILES = {
    "Cake-Cashloan-Q42025.xlsx": "CAKE",
    "VDS-Cashloan-Q42025.xlsx": "VDS",
    "MWG-CL-Q425.xlsx": "MWG",
    "VNPAY-CL-Q425.xlsx": "VNPAY",
    "VNPOST-CL-Q425.xlsx": "VNPOST",
    "Zalopay-CL-Q425.xlsx": "ZALOPAY",
}
Q425_PD_FILES = {
    "Cake-Payday-Q425.xlsx": "CAKE",
    "VDS-Payday-Q42025.xlsx": "VDS",
}


UE_ROOT = RAW.parent / "Unit Economic"
# Per-channel cohort UE files. Each file's Leadsheet/SAP_Leadsheet cached for that channel's sub-segment.
# Q1.26 = combined files only (dropdown locked to Cake) — only CAKE extractable.
COHORT_UE_FILES = {
    "CL": {
        "Q3.25": {
            "CAKE": "Q325/Cash Loan/Cake-Cashloan-Q32025.xlsx",
            "MWG": "Q325/Cash Loan/MWG-Cashloan-Q32025.xlsx",
            "VDS": "Q325/Cash Loan/VDS-Cashloan-Q32025.xlsx",
            "VNPAY": "Q325/Cash Loan/VNPAY-Cashloan-Q32025.xlsx",
            "ZALOPAY": "Q325/Cash Loan/Zalopay-Cashloan-Q32025.xlsx",
        },
        "Q4.25": {
            "CAKE": "Q425/Cash Loan/Cake-Cashloan-Q42025.xlsx",
            "MWG": "Q425/Cash Loan/MWG-CL-Q425.xlsx",
            "VDS": "Q425/Cash Loan/VDS-Cashloan-Q42025.xlsx",
            "VNPAY": "Q425/Cash Loan/VNPAY-CL-Q425.xlsx",
            "VNPOST": "Q425/Cash Loan/VNPOST-CL-Q425.xlsx",
            "ZALOPAY": "Q425/Cash Loan/Zalopay-CL-Q425.xlsx",
        },
        # Q1.26: extracted via Excel automation → q126_cohort_ue_cache.json (all channels)
    },
    "PD": {
        "Q3.25": {
            "CAKE": "Q325/Payday/Cake-payday-Q32025.xlsx",
            "VNPAY": "Q325/Payday/Vnpay-payday-Q32025.xlsx",
        },
        "Q4.25": {
            "CAKE": "Q425/Payday/Cake-Payday-Q425.xlsx",
            "VDS": "Q425/Payday/VDS-Payday-Q42025.xlsx",
        },
        # Q1.26: from cache
    },
    "PL": {
        "Q3.25": {
            "BE": "Q325/Paylater/Be-paylater-Q32025.xlsx",
            "MWG": "Q325/Paylater/MWG-paylater-q32025.xlsx",
            "VDS": "Q325/Paylater/VDS-Paylater-q32025.xlsx",
            "VNPAY": "Q325/Paylater/Vnpay-paylater-q32025.xlsx",
        },
        "Q4.25": {
            "BE": "Q425/Paylater/BE-Paylater-Q42025.xlsx",
            "MWG": "Q425/Paylater/MWG-Paylater-Q42025.xlsx",
            "VDS": "Q425/Paylater/VDS-Paylater-Q42025.xlsx",
            "VNPAY": "Q425/Paylater/VNPAY-Paylater-Q42025.xlsx",
        },
    },
}
# Cohort UE metric labels to search for in Leadsheet/SAP_Leadsheet.
# LOAN AMOUNT (CL/PD) and SPENDING (PL) = disbursement-per-loan.
COHORT_UE_LABELS = ["ANR", "TOI", "PROVISION", "OPEX", "PBT", "LOAN AMOUNT", "SPENDING"]


def _cohort_entry(disb, anr, toi, prov, opex, pbt, net_yield):
    """Build a cohort UE entry dict with %ANR + %Disbursement ratios."""
    return {
        "disbursement_mvnd": disb,
        "anr_mvnd": anr,
        "toi_mvnd": toi,
        "provision_mvnd": prov,
        "opex_mvnd": opex,
        "pbt_mvnd": pbt,
        "net_yield_mvnd": net_yield,
        # %ANR ratios
        "toi_per_anr_pct": (toi/anr*100) if (anr and toi is not None) else None,
        "provision_per_anr_pct": (prov/anr*100) if (anr and prov is not None) else None,
        "net_yield_per_anr_pct": (net_yield/anr*100) if (anr and net_yield is not None) else None,
        "pbt_per_anr_pct": (pbt/anr*100) if (anr and pbt is not None) else None,
        # %Disbursement ratios
        "toi_per_disb_pct": (toi/disb*100) if (disb and toi is not None) else None,
        "provision_per_disb_pct": (prov/disb*100) if (disb and prov is not None) else None,
        "net_yield_per_disb_pct": (net_yield/disb*100) if (disb and net_yield is not None) else None,
        "pbt_per_disb_pct": (pbt/disb*100) if (disb and pbt is not None) else None,
    }


def load_cohort_ue_per_channel():
    """Extract per-channel cohort UE (lifetime per-loan, mVND) from individual channel files.
    Returns {product: {quarter: {channel: {metric: lifetime_value}}}}.
    Robust: tries Leadsheet (CL/PD) then SAP_Leadsheet (PL); finds metric rows by label match;
    lifetime value = first numeric cell to the right of the label within 5 cols."""
    out = {}
    for product, quarters in COHORT_UE_FILES.items():
        out[product] = {}
        for quarter, channels in quarters.items():
            out[product][quarter] = {}
            for channel, relpath in channels.items():
                fpath = UE_ROOT / relpath
                if not fpath.exists():
                    continue
                try:
                    wb = openpyxl.load_workbook(fpath, data_only=True)
                except Exception:
                    continue
                ws = None
                for sheet_name in ["Leadsheet", "SAP_Leadsheet"]:
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        break
                if ws is None:
                    wb.close()
                    continue
                metrics = {}
                # Scan for metric labels in cols 1-13, take lifetime value (first numeric to the right)
                for r in range(1, min(ws.max_row + 1, 200)):
                    for c in range(1, 14):
                        v = ws.cell(r, c).value
                        if v and isinstance(v, str):
                            label = v.strip().upper()
                            if label in COHORT_UE_LABELS and label not in metrics:
                                # lifetime value = first numeric within next 5 cols
                                for cc in range(c + 1, c + 6):
                                    cell = ws.cell(r, cc).value
                                    if isinstance(cell, (int, float)):
                                        # normalize provision/opex to abs
                                        metrics[label] = abs(cell) if label in ("PROVISION", "OPEX") else cell
                                        break
                wb.close()
                if metrics:
                    anr = metrics.get("ANR")
                    toi = metrics.get("TOI")
                    prov = metrics.get("PROVISION")
                    pbt = metrics.get("PBT")
                    disb = metrics.get("LOAN AMOUNT") or metrics.get("SPENDING")  # disbursement-per-loan
                    net_yield = (toi - prov) if (toi is not None and prov is not None) else None
                    out[product][quarter][channel] = _cohort_entry(disb, anr, toi, prov, metrics.get("OPEX"), pbt, net_yield)

    # Merge Q1.26 per-channel cohort UE from Excel-automation cache (q126_cohort_ue_cache.json)
    cache_path = Path(__file__).resolve().parent / "q126_cohort_ue_cache.json"
    if cache_path.exists():
        import json as _json
        cache = _json.loads(cache_path.read_text())
        for product in ["CL", "PD", "PL"]:
            if product not in cache:
                continue
            out.setdefault(product, {})
            out[product]["Q1.26"] = {}
            for channel, m in cache[product].items():
                disb = m.get("disbursement")
                anr = m.get("anr")
                toi = m.get("toi")
                prov = abs(m["provision"]) if m.get("provision") is not None else None
                opex = abs(m["opex"]) if m.get("opex") is not None else None
                pbt = m.get("pbt")
                net_yield = (toi - prov) if (toi is not None and prov is not None) else None
                out[product]["Q1.26"][channel] = _cohort_entry(disb, anr, toi, prov, opex, pbt, net_yield)
    return out


def load_channel_sharing_params():
    """Extract per-channel partner sharing params from Q425 channel files' Leadsheet.
    Returns {product: {channel: {param: value}}}. product in {CL, PD}."""
    out = {"CL": {}, "PD": {}}
    for product, (folder, files) in [("CL", (Q425_CL_DIR, Q425_CL_FILES)), ("PD", (Q425_PD_DIR, Q425_PD_FILES))]:
        for fname, channel in files.items():
            fpath = folder / fname
            if not fpath.exists():
                continue
            try:
                wb = openpyxl.load_workbook(fpath, data_only=True)
                ws = wb["Leadsheet"]
                params = {}
                for param, row in SHARING_PARAM_ROWS.items():
                    v = ws.cell(row, 5).value
                    params[param] = v if isinstance(v, (int, float)) else None
                wb.close()
                out[product][channel] = params
            except Exception:
                continue
    return out


def overall_pbt_alloc_context(overall):
    """Returns {period: (pbo_overall, pbt_overall, disb_overall)} for PBT allocation.
    Allocation rule (theo Đạm): pbt_entity = pbo_entity − (pbo_overall − pbt_overall) × (disb_entity / disb_overall).
    Periods cover all section-4 + section-5.2 quarters."""
    def s(arr, idx):
        return sum(arr[i] for i in idx if arr[i] is not None)
    A25, A26, B26 = overall["2025 Actual"], overall["2026 Actual"], overall["2026 Budget"]
    return {
        "2025 FY":              (s(A25["pbo"], range(12)), s(A25["pbt"], range(12)), s(A25["disb"], range(12))),
        "2026 YTD":         (s(A26["pbo"], range(ACTUAL_2026_COMPLETE_MONTHS)),  s(A26["pbt"], range(ACTUAL_2026_COMPLETE_MONTHS)),  s(A26["disb"], range(ACTUAL_2026_COMPLETE_MONTHS))),
        "2026 Budget YTD":  (s(B26["pbo"], range(ACTUAL_2026_COMPLETE_MONTHS)),  s(B26["pbt"], range(ACTUAL_2026_COMPLETE_MONTHS)),  s(B26["disb"], range(ACTUAL_2026_COMPLETE_MONTHS))),
        "2026 Budget FY":       (s(B26["pbo"], range(12)), s(B26["pbt"], range(12)), s(B26["disb"], range(12))),
        "Q3.25": (s(A25["pbo"], [6,7,8]),    s(A25["pbt"], [6,7,8]),    s(A25["disb"], [6,7,8])),
        "Q4.25": (s(A25["pbo"], [9,10,11]),  s(A25["pbt"], [9,10,11]),  s(A25["disb"], [9,10,11])),
        "Q1.26": (s(A26["pbo"], [0,1,2]),    s(A26["pbt"], [0,1,2]),    s(A26["disb"], [0,1,2])),
    }


def allocate_pbt(pbo_entity, disb_entity, period_ctx):
    if pbo_entity is None or disb_entity is None or not period_ctx:
        return None
    pbo_ov, pbt_ov, disb_ov = period_ctx
    if not disb_ov:
        return None
    overhead = pbo_ov - pbt_ov  # segment overhead (positive when PBO > PBT)
    share = disb_entity / disb_ov
    return pbo_entity - overhead * share


def apply_pbt_allocation(overall, deep, by_prod_q):
    """Inject pbt_allocated_by_disb_share into section 4 + 5.2 entries."""
    ctx = overall_pbt_alloc_context(overall)
    # Section 4: deep[ch][period][product]
    for ch in deep:
        for period, products in deep[ch].items():
            period_ctx = ctx.get(period)
            for product, m in products.items():
                m["pbt_allocated_by_disb_share"] = allocate_pbt(m.get("pbo"), m.get("disb"), period_ctx)
    # Section 5.2: by_prod_q[product][quarter]
    for product in by_prod_q:
        for quarter, m in by_prod_q[product].items():
            period_ctx = ctx.get(quarter)
            m["pbt_allocated_by_disb_share"] = allocate_pbt(m.get("pbo"), m.get("disb"), period_ctx)


def build_comprehensive_matrix(deep, channel_matrix_2025, q126_cl_inputs, q126_pd_inputs, paylater_ch):
    """Comprehensive (period, product, channel) flat matrix combining ALL metrics for team queries.

    Periods: '2025 FY', '2026 YTD', '2026 Budget YTD', '2026 Budget FY'.
    Channels: CAKE, VDS (incl. merged VDS-PR), ZLP, MWG (full P&L+risk data) + others (CL: VNPAY, BE, VNPOST, MISA;
              PD: VNPAY, ZALOPAY; PL: Total, VDS, VDS-EPASS, BE, VNPAY, FPT, MWG) where data available.
    Products: CL, OD, PD, PL.
    Metric set (None where data missing):
      P&L: disb, toi, provision, pbo (Tỷ VND, flow)
      Risk: balance (Tỷ VND, state EOP), npl (%), lg2 (%)
      Pricing inputs (Q1.26 CL/PD only): ir, cof, nim, avg_ticket, avg_tenor, approval_rate, signed_rate, insurance_pen
      Paylater specific: accumulated_account, active_account, active_rate, new_issued_q1,
                         spending_per_active_mar, avg_balance_per_active
    """
    matrix = {p: {} for p in ["2025 FY", "2026 YTD", "2026 Budget YTD", "2026 Budget FY"]}

    # Channels in deep (CAKE/VDS/ZLP/MWG × CL/OD/Payday) — VDS-PR đã merge vào VDS
    for period in ["2025 FY", "2026 YTD", "2026 Budget YTD", "2026 Budget FY"]:
        for channel in GROUP_CHANNELS:
            if channel not in deep or period not in deep[channel]:
                continue
            for product, metrics in deep[channel][period].items():
                key = f"{product}_{channel}"
                entry = {
                    "product": product, "channel": channel, "period": period,
                    "disb_bil_vnd": metrics.get("disb"),
                    "toi_bil_vnd": metrics.get("toi"),
                    "provision_bil_vnd": metrics.get("prov"),
                    "pbo_bil_vnd": metrics.get("pbo"),
                    "pbt_allocated_bil_vnd": metrics.get("pbt_allocated_by_disb_share"),
                    "balance_bil_vnd": metrics.get("balance"),
                    "npl_cake": metrics.get("npl_cake"),
                    "npl_cic": metrics.get("npl_cic"),
                    "lg2": metrics.get("lg2"),
                }
                # Compute ratios
                disb = entry["disb_bil_vnd"]; toi = entry["toi_bil_vnd"]; prov = entry["provision_bil_vnd"]
                pbo = entry["pbo_bil_vnd"]; bal = entry["balance_bil_vnd"]
                entry["yield_toi_per_disb_pct"] = (toi/disb*100) if (disb and toi is not None) else None
                entry["cor_provision_per_disb_pct"] = (prov/disb*100) if (disb and prov is not None) else None
                entry["pbo_margin_per_disb_pct"] = (pbo/disb*100) if (disb and pbo is not None) else None
                entry["yield_toi_per_anr_pct"] = (toi/bal*100) if (bal and toi is not None) else None
                entry["cor_provision_per_anr_pct"] = (prov/bal*100) if (bal and prov is not None) else None
                entry["pbo_per_anr_pct"] = (pbo/bal*100) if (bal and pbo is not None) else None
                matrix[period][key] = entry

    # Add Q1.26 pricing inputs to 2026 YTD CL channels (Note: pricing is Q1 from DATA sheet, ~= YTD proxy)
    period_inputs = "2026 YTD"
    for channel, d in q126_cl_inputs.items():
        if channel == "All":
            continue
        # Map channel name (Q126 inputs may use diff casing)
        ch_norm = {"Cake": "CAKE", "VDS": "VDS"}.get(channel, channel)
        key = f"CL_{ch_norm}"
        if key not in matrix[period_inputs]:
            # Channel has only inputs (no P&L deep data) — create entry
            matrix[period_inputs][key] = {"product": "CL", "channel": ch_norm, "period": period_inputs}
        ir = d.get("ir") or 0
        cof = d.get("cof") or 0
        matrix[period_inputs][key].update({
            "ir_pct": ir,
            "cof_pct": cof,
            "nim_pct": (ir - cof) if (ir and cof) else None,
            "avg_ticket_mvnd": d.get("avg_ticket"),
            "avg_tenor_months": d.get("avg_tenor"),
            "approval_rate": d.get("approval_rate"),
            "signed_rate": d.get("signed_rate"),
            "insurance_penetration": d.get("insurance_pen"),
        })

    for channel, d in q126_pd_inputs.items():
        if channel == "All":
            continue
        ch_norm = {"Cake": "CAKE", "VDS": "VDS"}.get(channel, channel)
        key = f"PD_{ch_norm}"  # PD = Payday
        # Map PD product name to "Payday" used in deep
        key_alt = f"Payday_{ch_norm}"
        target_key = key_alt if key_alt in matrix[period_inputs] else key
        if target_key not in matrix[period_inputs]:
            matrix[period_inputs][target_key] = {"product": "Payday", "channel": ch_norm, "period": period_inputs}
        ir = d.get("ir") or 0
        cof = d.get("cof") or 0
        matrix[period_inputs][target_key].update({
            "ir_pct": ir,
            "cof_pct": cof,
            "nim_pct": (ir - cof) if (ir and cof) else None,
            "avg_ticket_mvnd": d.get("avg_ticket"),
            "avg_tenor_months": d.get("avg_tenor"),
            "approval_rate": d.get("approval_rate"),
            "signed_rate": d.get("signed_rate"),
            "insurance_penetration": d.get("insurance_pen"),
        })

    # Paylater channels — add to 2026 YTD period (Q1 data closest)
    for channel, d in paylater_ch.items():
        if channel == "Total":
            continue
        key = f"PL_{channel}"
        matrix[period_inputs][key] = {
            "product": "PL", "channel": channel, "period": period_inputs,
            "disb_bil_vnd": (d.get("spending") or 0) / 1e9 if d.get("spending") else None,  # spending = disb-equiv
            "toi_bil_vnd": (d.get("toi") or 0) / 1e9 if d.get("toi") else None,
            "provision_bil_vnd": (d.get("provision") or 0) / 1e9 if d.get("provision") else None,
            "pbo_bil_vnd": (d.get("pbo") or 0) / 1e9 if d.get("pbo") else None,
            "balance_bil_vnd": (d.get("balance") or 0) / 1e9 if d.get("balance") else None,
            "npl_cake": d.get("npl_cake"),
            "npl_cic": d.get("npl_cic"),
            "lg2": d.get("lg2"),
            # Paylater specific
            "accumulated_account": d.get("accumulated_account"),
            "active_account": d.get("active_account"),
            "active_rate": d.get("active_rate"),
            "new_issued_q1": d.get("new_issued_q1"),
            "spending_per_active_mar_vnd": d.get("spending_per_active_mar"),
            "avg_balance_per_active_vnd": d.get("avg_balance_per_active"),
        }
        # Compute ratios
        disb = matrix[period_inputs][key]["disb_bil_vnd"]
        toi = matrix[period_inputs][key]["toi_bil_vnd"]
        prov = matrix[period_inputs][key]["provision_bil_vnd"]
        pbo = matrix[period_inputs][key]["pbo_bil_vnd"]
        bal = matrix[period_inputs][key]["balance_bil_vnd"]
        matrix[period_inputs][key]["yield_toi_per_disb_pct"] = (toi/disb*100) if (disb and toi is not None) else None
        matrix[period_inputs][key]["cor_provision_per_disb_pct"] = (prov/disb*100) if (disb and prov is not None) else None
        matrix[period_inputs][key]["pbo_margin_per_disb_pct"] = (pbo/disb*100) if (disb and pbo is not None) else None
        matrix[period_inputs][key]["yield_toi_per_anr_pct"] = (toi/bal*100) if (bal and toi is not None) else None
        matrix[period_inputs][key]["cor_provision_per_anr_pct"] = (prov/bal*100) if (bal and prov is not None) else None
        matrix[period_inputs][key]["pbo_per_anr_pct"] = (pbo/bal*100) if (bal and pbo is not None) else None

    # Add 2025 FY broader channels from channel_matrix_2025 (BEG, NGS, MWG, VNPAY, ZALOPAY, VNPOST, VNPOST-PR, MISA)
    for channel, products_dict in channel_matrix_2025.items():
        for product, metrics in products_dict.items():
            # Map product name CL → CL, Overdraft → OD, Payday → Payday
            prod_norm = {"CashLoan": "CL", "Overdraft": "OD", "Payday": "Payday"}.get(product, product)
            key = f"{prod_norm}_{channel}"
            if key in matrix["2025 FY"]:
                continue  # already have from deep
            matrix["2025 FY"][key] = {
                "product": prod_norm, "channel": channel, "period": "2025 FY",
                "disb_bil_vnd": metrics.get("disb"),
                "toi_bil_vnd": metrics.get("toi"),
                "pbo_bil_vnd": metrics.get("pbo"),
            }
            disb = metrics.get("disb"); toi = metrics.get("toi"); pbo = metrics.get("pbo")
            matrix["2025 FY"][key]["yield_toi_per_disb_pct"] = (toi/disb*100) if (disb and toi is not None) else None
            matrix["2025 FY"][key]["pbo_margin_per_disb_pct"] = (pbo/disb*100) if (disb and pbo is not None) else None

    # Post-process: add Net Yield (TOI − Risk) to every entry that has both toi + provision.
    # Net Yield is the truer cross-channel comparison metric vì partner sharing makes raw TOI misleading.
    for period in matrix:
        for key, entry in matrix[period].items():
            toi = entry.get("toi_bil_vnd")
            prov = entry.get("provision_bil_vnd")
            disb = entry.get("disb_bil_vnd")
            bal = entry.get("balance_bil_vnd")
            if toi is not None and prov is not None:
                net_yield = toi - prov  # provision stored as positive expense magnitude
                entry["net_yield_toi_minus_risk_bil_vnd"] = net_yield
                entry["net_yield_per_disb_pct"] = (net_yield/disb*100) if disb else None
                entry["net_yield_per_anr_pct"] = (net_yield/bal*100) if bal else None
            else:
                entry["net_yield_toi_minus_risk_bil_vnd"] = None
                entry["net_yield_per_disb_pct"] = None
                entry["net_yield_per_anr_pct"] = None

    return matrix


def export_ground_truth_json(overall, by_product, channel_matrix, deep, ue_q34, ue_q1, by_prod_q, cl_ue, q126_cl_inputs, q126_pd_inputs=None, paylater_q1=None, paylater_ch=None, ch_monthly=None):
    """Export all extracted data to JSON for team-hub Claude Project ground truth."""
    import json
    from datetime import date
    out = {
        "_meta": {
            "generated_date": date.today().isoformat(),
            "currency": "Tỷ VND (Bil VND), unless noted",
            "scope": "Cake Lending segment (Head of Lending Product = Đạm)",
            "actuals_2026_complete_through_month": ACTUAL_2026_COMPLETE_MONTHS,
            "sources": {
                "actual_2025": "Raw/Finance/P&L/P&L_2025_final.xlsx",
                "actual_2026": "Raw/Finance/P&L/P&L_2026_Actual.xlsx",
                "budget_2026": "Raw/Finance/P&L/P&L-Budget-2026.xlsx",
                "ue_q325_q425": "Raw/Finance/Unit Economic/Q425/Unit Economics Q42025 Lending.xlsx",
                "ue_q126": "Raw/Finance/Unit Economic/Q126/*.xlsx",
            },
        },
        "section_1_overall_pnl": {
            "description": "Lending segment top-level P&L: monthly + YTD + FY for 3 periods (2025 Actual, 2026 Actual, 2026 Budget). Metrics: Disbursement, TOI, Provision, Profit Before Overhead (PBO), PBT.",
            "data": overall,
        },
        "section_2_by_product_pnl": {
            "description": "By product (CashLoan, Overdraft, Payday, Paylater) — Disb, TOI, PBO across 3 periods.",
            "data": by_product,
        },
        "section_3_channel_x_product_2025": {
            "description": "FY 2025 channel × product matrix (Disb, TOI, PBO).",
            "data": channel_matrix,
        },
        "section_4_channel_deep_dive": {
            "description": "Channel deep dive (CAKE / VDS / ZLP / MWG — VDS-PR đã merge vào VDS) × product (CL/OD/Payday) × period (2025 FY, 2026 YTD, 2026 Budget YTD, 2026 Budget FY). Metrics: disb, toi, prov, pbo, balance, npl_cake, npl_cic, lg2. Budget chỉ có cho CAKE + VDS.",
            "data": deep,
        },
        "section_4_1_channel_budget_monthly": {
            "description": "Monthly Actual vs Budget 2026 for CAKE + VDS channels × product (CL/Payday/OD). Structure: {channel: {product: {actual_monthly: {metric: [12 values]}, budget_monthly: {metric: [12 values]}}}}. Metrics: disb, toi, prov, pbo (Tỷ VND). Actual closed through month " + str(ACTUAL_2026_COMPLETE_MONTHS) + " (rest = null). VDS-PR merged into VDS. Budget = full 12 months. Dùng để xem pacing monthly, identify tháng nào miss/beat.",
            "data": ch_monthly or {},
        },
        "section_5_1_unit_economics_aggregate": {
            "description": "Lending aggregate Unit Economics: Q3.25 + Q4.25 from UE Q425 SUM file, Q1.26 derived from P&L Actual + Balance ENR. Metrics: Disb, Balance(ANR), TOI, Provision, OPEX, PBT, NPL, LG2.",
            "q3_25": ue_q34["Q3.25"],
            "q4_25": ue_q34["Q4.25"],
            "q1_26": ue_q1,
        },
        "section_5_2_by_product_quarterly_aggregate": {
            "description": "By product quarterly aggregate (Q3.25 + Q4.25 + Q1.26) Tỷ VND. Source: P&L files (2025 channels summed → product aggregate, 2026 product sheets).",
            "data": by_prod_q,
        },
        "section_5_2_1_cl_cohort_per_loan": {
            "description": "CL per-loan cohort UE (mVND/loan + %ANR lifetime ratio). Source: Q425 SUM file CL sheet. Channels: CL all, CAKE, VDS, MWG, VNPAY, ZLP. Q3.25 + Q4.25 only (Q1.26 cohort per channel not available — Leadsheet locked to Cake).",
            "data": cl_ue,
        },
        "section_5_3_q126_cl_channel_inputs": {
            "description": "Q1.26 CashLoan channel-level INPUTS (pricing/funnel) from CashLoan_Q1_26 DATA sheet R45-R53. Channels: BE/CAKE/MISA/MWG/VNPOST/VDS/VNPAY/ZALOPAY + All. Metrics: IR, COF, AVG Ticket size, AVG Tenor, Approval rate, Signed rate, Insurance rate, Insurance penetration.",
            "data": q126_cl_inputs,
        },
        "section_5_3_q126_pd_channel_inputs": {
            "description": "Q1.26 Payday channel-level INPUTS from Payday_Q1_26 DATA sheet R41-R45. Channels: VDS/CAKE/VNPAY/ZALOPAY + All. Same input metrics as CL.",
            "data": q126_pd_inputs or {},
        },
        "section_5_2_3_paylater_q126_metrics": {
            "description": "Paylater Q1.26 detailed product-specific metrics (Top-level, all-channel). Source: P&L_2026_Actual Paylater sheet. State metrics = end-of-Q1 (Mar 26). Flow = sum Jan+Feb+Mar. Includes: New issued (= Δ accumulated), Lũy kế accumulated account, Active account, % Active rate, Spending Q1, Balance Mar 26, Spending/active, Spending/accumulated, Avg balance per active (proxy for avg limit).",
            "data": paylater_q1 or {},
        },
        "section_5_2_4_paylater_channels_q126": {
            "description": "Paylater Q1.26 per-channel metrics. Channels: Total, VDS, VDS-EPASS, BE, VNPAY, FPT, MWG. Same metric set as 5.2.3 plus per-channel TOI/Provision/PBO. Source: Paylater sheet channel sub-sections (R54, R107, R160, R213, R266, R319).",
            "data": paylater_ch or {},
        },
        "section_6_comprehensive_channel_product_matrix": {
            "description": "MASTER QUERY TABLE — comprehensive channel × product matrix. Flat structure {period: {<product>_<channel>: {all_metrics}}}. Periods: 2025 FY, 2026 YTD (Jan through last closed month), 2026 Budget YTD, 2026 Budget FY. Each entry contains: product, channel, period, P&L absolutes (disb/toi/provision/pbo Tỷ VND), risk state (balance/npl/lg2), computed ratios (%Disb: yield/CoR/PBO_margin; %ANR: yield/CoR/PBO), NET YIELD (toi_minus_risk: net_yield_toi_minus_risk_bil_vnd + net_yield_per_disb_pct + net_yield_per_anr_pct), pricing inputs (ir/cof/nim/avg_ticket/avg_tenor/approval/signed/insurance_pen — Q1.26 CL+PD only), Paylater-specific (accumulated/active/active_rate/new_issued/spending_per_active/avg_balance_per_active). None where data missing. Use this section first cho mọi câu hỏi về channel × product breakdown.",
            "data": build_comprehensive_matrix(deep, channel_matrix, q126_cl_inputs, q126_pd_inputs or {}, paylater_ch or {}),
        },
        "section_7_partner_sharing_params": {
            "description": "Partner revenue/risk sharing params per channel (CL + PD). Source: Q425 channel files' Leadsheet R47-R54. CRITICAL CONTEXT: mỗi partner có cấu trúc sharing KHÁC NHAU → raw TOI comparison giữa các channel bị misleading. Phải dùng Net Yield (TOI − Risk) trong Section 6 để so sánh fair. Params: sharing_toi_risk_dpd90 (% TOI+Risk shared with partner from DPD90+), sharing_toi_risk_dpd180 (same from DPD180+), sharing_cii_only / sharing_cof_only / sharing_nfi_only / sharing_risk_dpd90_only / sharing_risk_dpd180_only (component-specific sharing). Vd: VDS share 50% TOI+Risk từ DPD180+, MWG share 35% từ DPD90+, ZALOPAY không share. None = không share / Cake giữ 100%.",
            "data": load_channel_sharing_params(),
        },
        "section_8_cohort_ue_per_channel": {
            "description": "Per-channel COHORT Unit Economics (lifetime per-loan, mVND). Source: individual channel files trong Q325 + Q425 folders (mỗi file Leadsheet/SAP_Leadsheet đã cache cohort UE của đúng channel đó). Structure: {product: {quarter: {channel: {anr_mvnd, toi_mvnd, provision_mvnd, opex_mvnd, pbt_mvnd, toi_per_anr_pct, provision_per_anr_pct, pbt_per_anr_pct, net_yield_mvnd}}}}. Coverage: CL (CAKE/MWG/VDS/VNPAY/ZALOPAY/VNPOST), PD (CAKE/VNPAY/VDS), PL (BE/MWG/VDS/VNPAY) cho Q3.25 + Q4.25. ⚠️ Q1.26 = CHỈ có CAKE (file Q126 là combined file, sub-segment dropdown khóa ở Cake) — muốn channel khác Q1.26 phải mở file, đổi dropdown sub-segment, save riêng từng file. Đây là per-LOAN lifetime view (mVND/loan), khác với aggregate Tỷ VND ở section 4/6.",
            "data": load_cohort_ue_per_channel(),
        },
    }
    # Carry forward quarterly snapshots the monthly generator can't reproduce.
    # section_9 = per-loan P&L+pricing Q1.26 extracted via AppleScript (refresh-q126-cohort.md),
    # not built from the monthly Excel feed. It only changes on a new QUARTER, so a monthly
    # run must preserve it from the prior canonical bundle instead of dropping it.
    prior = Path(__file__).resolve().parent / "team-hub" / "lending-pl-data.json"
    for carry_key in ("section_9_ue_q126_per_channel_extracted",):
        if carry_key in out:
            continue
        if prior.exists():
            prior_data = json.loads(prior.read_text())
            if carry_key in prior_data:
                out[carry_key] = prior_data[carry_key]
                print(f"  ↪ carried forward {carry_key} from prior bundle "
                      f"(extracted_at {prior_data[carry_key].get('extracted_at', '?')})")
            else:
                print(f"  ⚠ {carry_key} not in prior bundle — skipped")
        else:
            print(f"  ⚠ no prior bundle at {prior} — {carry_key} not carried")

    out_path = Path("/tmp/lending-pl-data.json")
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2, default=str))
    print(f"✓ Ground-truth JSON written: {out_path}")
    return out_path


def render_by_product_year(by_product, period_key, period_label, ytd_only=False):
    """Simple by-product table for a single period. ytd_only=True → sum first 4 months."""
    metrics = [("disb", "Disbursement"), ("toi", "TOI"), ("pbo", "Profit before overhead")]
    parts = []
    parts.append("<table class='data'><thead><tr><th>Product</th>" +
                 "".join(f"<th>{lbl}</th>" for _, lbl in metrics) + "</tr></thead><tbody>")
    for product in ["CashLoan", "Overdraft", "Payday", "Paylater"]:
        parts.append(f"<tr><td class='label'>{product}</td>")
        for mkey, _ in metrics:
            arr = by_product[product][mkey].get(period_key, [None]*12)
            val = safe_sum(arr[:ACTUAL_2026_COMPLETE_MONTHS]) if ytd_only else safe_sum(arr)
            parts.append(f"<td>{fmt(val)}</td>")
        parts.append("</tr>")
    parts.append("</tbody></table>")
    return "\n".join(parts)


def render_tab_pnl_2025(overall, channel_matrix, by_product):
    parts = []
    parts.append("<h2>1.1 — Overall Lending P&L 2025 (monthly)</h2>")
    parts.append(render_overall_table(overall, ["2025 Actual"], show_delta=False))
    parts.append("<h2>1.2 — By Product 2025 FY</h2>")
    parts.append("<p class='legend'>Full-year 2025 actual per product. PBT chỉ có ở top-level (overhead allocated to segment).</p>")
    parts.append(render_by_product_year(by_product, "2025 Actual", "2025 FY", ytd_only=False))
    parts.append("<h2>1.3 — Channel × Product Matrix 2025 FY</h2>")
    parts.append(render_by_channel(channel_matrix))
    return "\n".join(parts)


def render_tab_pnl_2026(overall, by_product, deep, ch_monthly=None):
    parts = []
    parts.append(f"<h2>2.1 — Overall Lending P&L 2026 (Actual {YTD_LABEL} vs Budget)</h2>")
    parts.append(render_overall_table(overall, ["2026 Actual", "2026 Budget"], show_delta=True))
    parts.append("<h2>2.2 — By Product 2026 (Actual YTD vs Budget vs YoY)</h2>")
    parts.append(f"<p class='legend'>So sánh 2026 Actual {YTD_LABEL} với Budget YTD và YoY vs 2025. PBT chỉ có top-level.</p>")
    parts.append(render_by_product(by_product))
    parts.append("<h2>2.3 — Channel Deep Dive (CAKE / VDS / Others)</h2>")
    parts.append(render_channel_deep_dive(deep, overall))
    if ch_monthly:
        parts.append(render_channel_monthly_budget(ch_monthly))
    return "\n".join(parts)


def render_tab_unit_econ(overall):
    parts = []
    parts.append("<h2>3 — Unit Economics — Quarterly Trend (Q3.25 → Q4.25 → Q1.26)</h2>")
    parts.append(render_unit_economics(overall))
    return "\n".join(parts)


def main():
    global ACTUAL_2026_COMPLETE_MONTHS, YTD_LABEL, ANNUALIZE
    ACTUAL_2026_COMPLETE_MONTHS = detect_complete_months()
    YTD_LABEL = f"YTD-{MONTHS[ACTUAL_2026_COMPLETE_MONTHS - 1]}"
    ANNUALIZE = 12 / ACTUAL_2026_COMPLETE_MONTHS
    print(f"→ 2026 Actual closed through month {ACTUAL_2026_COMPLETE_MONTHS} "
          f"({YTD_LABEL}); annualize ×{ANNUALIZE:.2f}")

    overall = load_overall()
    by_product = load_by_product()
    channel_matrix = load_channel_matrix_2025()
    deep = load_channel_deep()
    ch_monthly = load_channel_budget_monthly()

    # Load UE/cohort/quarterly data
    ue_q34 = load_ue_q3q4_25()
    ue_q1 = derive_ue_q1_26(overall)
    by_prod_q = load_by_product_quarterly()
    cl_ue = load_cl_per_loan_ue()
    q126_cl_inputs = load_q126_cl_channel_inputs()
    q126_pd_inputs = load_q126_pd_channel_inputs()
    paylater_q1 = load_paylater_q126_metrics()
    paylater_ch = load_paylater_channel_metrics()

    # PBT allocation by Disbursement share (segment overhead → product/channel) — apply BEFORE rendering
    apply_pbt_allocation(overall, deep, by_prod_q)

    html_doc = HTML_TEMPLATE.format(
        tab_pnl_2025=render_tab_pnl_2025(overall, channel_matrix, by_product),
        tab_pnl_2026=render_tab_pnl_2026(overall, by_product, deep, ch_monthly),
        tab_unit_econ=render_tab_unit_econ(overall),
    )

    export_ground_truth_json(overall, by_product, channel_matrix, deep, ue_q34, ue_q1, by_prod_q, cl_ue, q126_cl_inputs, q126_pd_inputs, paylater_q1, paylater_ch, ch_monthly)

    OUT.write_text(html_doc, encoding="utf-8")
    print(f"✓ Report written: {OUT}")
    print(f"  open {OUT}")


if __name__ == "__main__":
    main()
