#!/usr/bin/env python3
"""Lending P&L Performance Tracker — Cake.

So sánh Disbursement, TOI, Profit before overhead, PBT giữa:
  - Actual 2025 (FY, monthly)
  - Actual 2026 (YTD, monthly)
  - Budget 2026 (FY, monthly)

Source files trong /Raw/Finance/P&L/:
  - P&L_2025_final.xlsx    sheet=Lending  (raw VND)
  - P&L_2026_Actual.xlsx   sheet=Total    (raw VND)
  - P&L-Budget-2026.xlsx   sheet=Lending  (already Bil VND)

Usage:
  python3 lending_pl_tracker.py                # console table
  python3 lending_pl_tracker.py --md           # markdown table
  python3 lending_pl_tracker.py --ytd-months 5 # compare YTD up to May (default 4)

Nếu Finance team restructure file, script sẽ FAIL LOUD với message
"row X col Y expected '...' got '...'" — update ROW_MAP và rerun.
"""
import argparse
import openpyxl
from pathlib import Path
import sys

REPO_ROOT = Path(__file__).resolve().parents[2]
RAW = REPO_ROOT / "Raw" / "Finance" / "P&L"

# (row, label_col, expected_label_exact_after_strip_lower)
CONFIG = {
    "actual_2025": {
        "path": RAW / "P&L_2025_final.xlsx",
        "sheet": "Lending",
        "month_start": 16, "month_end": 27,   # Jan-Dec 2025
        "unit_divisor": 1e9,
        "rows": {
            "disb": (7,  12, "Disbursement"),
            "toi":  (19, 12, "TOI"),
            "pbo":  (46, 12, "Profit before allocation cost"),
            "pbt":  (45, 12, "Profit after allocation cost"),
        },
    },
    "actual_2026": {
        "path": RAW / "P&L_2026_Actual.xlsx",
        "sheet": "Total",
        "month_start": 6, "month_end": 17,    # Jan-Dec 2026 (Apr+ may be empty)
        "unit_divisor": 1e9,
        "rows": {
            "disb": (6,  2, "Disbursement"),
            "toi":  (17, 2, "TOI"),
            "pbo":  (52, 2, "Profit before allocation cost"),
            "pbt":  (53, 2, "Profit after allocation cost"),
        },
    },
    "budget_2026": {
        "path": RAW / "P&L-Budget-2026.xlsx",
        "sheet": "Lending",
        "month_start": 62, "month_end": 73,   # Jan-Dec 2026
        "unit_divisor": 1.0,
        "rows": {
            "disb": (7,  6, "disburse/spend"),
            "toi":  (11, 6, "TOI"),
            "pbo":  (45, 6, "profit_before_overhead"),
            "pbt":  (47, 6, "profit_after_overhead"),
        },
    },
}

METRIC_LABELS = {
    "disb": "Disbursement (Giải ngân)",
    "toi":  "TOI",
    "pbo":  "Profit before overhead",
    "pbt":  "PBT (Profit after allocation)",
}

MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]


def load_dataset(name, cfg):
    wb = openpyxl.load_workbook(cfg["path"], data_only=True)
    ws = wb[cfg["sheet"]]
    out = {}
    for metric, (row, label_col, expected) in cfg["rows"].items():
        actual = ws.cell(row, label_col).value
        if actual is None or str(actual).strip().lower() != expected.strip().lower():
            raise ValueError(
                f"[{name}] R{row} C{label_col}: expected label '{expected}' but got "
                f"'{actual}'. File structure changed — update CONFIG in lending_pl_tracker.py"
            )
        vals = []
        for c in range(cfg["month_start"], cfg["month_end"] + 1):
            v = ws.cell(row, c).value
            vals.append(None if (v is None or isinstance(v, str)) else v / cfg["unit_divisor"])
        out[metric] = vals
    wb.close()
    return out


def fmt(v, width=8):
    return f"{'-':>{width}}" if v is None else f"{v:>{width},.1f}"


def sum_n(arr, n=None):
    pool = arr if n is None else arr[:n]
    vals = [v for v in pool if v is not None]
    return sum(vals) if vals else None


def console_output(a25, a26, b26, ytd_n):
    L = []
    L.append("=" * 140)
    L.append(f"LENDING P&L PERFORMANCE TRACKING — Đơn vị: Tỷ VND  ·  YTD = first {ytd_n} months ({', '.join(MONTHS[:ytd_n])})")
    L.append("=" * 140)
    for metric, label in METRIC_LABELS.items():
        L.append(f"\n### {label}")
        L.append(f"  {'Period':14} | " + " ".join(f"{m:>8}" for m in MONTHS) + f" | {'YTD':>9} {'FY':>10}")
        for tag, data in [("2025 Actual", a25[metric]), ("2026 Actual", a26[metric]), ("2026 Budget", b26[metric])]:
            L.append(f"  {tag:14} | {' '.join(fmt(v) for v in data)} | {fmt(sum_n(data, ytd_n), 9)} {fmt(sum_n(data), 10)}")
        a_ytd, b_ytd, p_ytd, b_fy = sum_n(a26[metric], ytd_n), sum_n(b26[metric], ytd_n), sum_n(a25[metric], ytd_n), sum_n(b26[metric])
        if a_ytd is not None and b_ytd:
            L.append(f"  Δ vs BG YTD : {a_ytd - b_ytd:+,.1f} Bil ({(a_ytd - b_ytd)/b_ytd*100:+.0f}%)")
        if a_ytd is not None and p_ytd:
            L.append(f"  YoY vs 2025: {a_ytd - p_ytd:+,.1f} Bil ({(a_ytd - p_ytd)/p_ytd*100:+.0f}%)")
        if a_ytd is not None and b_fy:
            L.append(f"  Pace        : {a_ytd:,.1f} / {b_fy:,.1f} = {a_ytd/b_fy*100:.1f}% FY done")
    return "\n".join(L)


def md_output(a25, a26, b26, ytd_n):
    L = []
    L.append("# Lending P&L Performance Tracking")
    L.append("")
    L.append(f"**Đơn vị: Tỷ VND** · YTD = sum {', '.join(MONTHS[:ytd_n])} · Generated by `Projects/Finance-Tracking/lending_pl_tracker.py`")
    L.append("")
    L.append("| Metric | 2025 Actual FY | 2026 Actual YTD | 2026 Budget YTD | Δ vs BG | YoY | 2026 Budget FY | % FY done |")
    L.append("|---|---:|---:|---:|---:|---:|---:|---:|")
    for metric, label in METRIC_LABELS.items():
        a25_fy = sum_n(a25[metric])
        a_ytd = sum_n(a26[metric], ytd_n)
        b_ytd = sum_n(b26[metric], ytd_n)
        b_fy = sum_n(b26[metric])
        p_ytd = sum_n(a25[metric], ytd_n)
        var = (a_ytd - b_ytd) if (a_ytd is not None and b_ytd is not None) else None
        var_pct = (var/b_ytd*100) if (var is not None and b_ytd) else None
        yoy = (a_ytd - p_ytd) if (a_ytd is not None and p_ytd is not None) else None
        yoy_pct = (yoy/p_ytd*100) if (yoy is not None and p_ytd) else None
        pace = (a_ytd/b_fy*100) if (a_ytd is not None and b_fy) else None
        def s(v): return f"{v:,.1f}" if v is not None else "-"
        var_c = f"{s(var)} ({var_pct:+.0f}%)" if var is not None else "-"
        yoy_c = f"{s(yoy)} ({yoy_pct:+.0f}%)" if yoy is not None else "-"
        pace_c = f"{pace:.1f}%" if pace is not None else "-"
        L.append(f"| **{label}** | {s(a25_fy)} | {s(a_ytd)} | {s(b_ytd)} | {var_c} | {yoy_c} | {s(b_fy)} | {pace_c} |")
    L.append("")
    L.append("## Monthly breakdown — 2026 Actual vs Budget")
    L.append("")
    header = "| | " + " | ".join(MONTHS) + " |"
    sep = "|---|" + "|".join(["---:"] * 12) + "|"
    L.append(header); L.append(sep)
    for metric, label in METRIC_LABELS.items():
        L.append(f"| **{label}** Actual | " + " | ".join(fmt(v).strip() for v in a26[metric]) + " |")
        L.append(f"| {label} Budget | " + " | ".join(fmt(v).strip() for v in b26[metric]) + " |")
        deltas = [(a26[metric][i] - b26[metric][i]) if (a26[metric][i] is not None and b26[metric][i] is not None) else None for i in range(12)]
        L.append(f"| Δ vs BG | " + " | ".join((f"{v:+,.1f}" if v is not None else "-") for v in deltas) + " |")
    return "\n".join(L)


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--md", action="store_true", help="Markdown output")
    p.add_argument("--ytd-months", type=int, default=4, help="YTD month count (default 4)")
    args = p.parse_args()

    a25 = load_dataset("actual_2025", CONFIG["actual_2025"])
    a26 = load_dataset("actual_2026", CONFIG["actual_2026"])
    b26 = load_dataset("budget_2026", CONFIG["budget_2026"])

    print(md_output(a25, a26, b26, args.ytd_months) if args.md else console_output(a25, a26, b26, args.ytd_months))


if __name__ == "__main__":
    main()
